"""
图书管理系统后台主界面
"""

import os
import sqlite3
import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
from tkinter.filedialog import askopenfilename, asksaveasfilename

import chineseize_matplotlib
import matplotlib
matplotlib.use('TkAgg')  # 指定后端为TkAgg
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import pandas as pd
from openpyxl import Workbook, load_workbook


def _project_root():
    # code 目录的上一级即项目根目录
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def _user_data_dir():
    user_data_dir = os.path.join(_project_root(), "user_data")
    os.makedirs(user_data_dir, exist_ok=True)
    return user_data_dir


def _book_db_path():
    # 图书数据库（与用户数据库分离）
    return os.path.join(_user_data_dir(), "book_info.db")


def _user_db_path():
    # 用户数据库
    return os.path.join(_user_data_dir(), "user_info.db")


def load_all_users():
    """读取 user 表所有用户，返回 [[username, password, phone, role], ...]。"""
    conn = sqlite3.connect(_user_db_path())
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT username, password, phone, role FROM user")
        rows = cursor.fetchall()
    finally:
        conn.close()
    return [list(row) for row in rows]


def add_user(username, password, phone, role='student'):
    """添加新用户。"""
    conn = sqlite3.connect(_user_db_path())
    try:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT OR IGNORE INTO user (username, password, phone, role) VALUES (?, ?, ?, ?)",
            (username, password, phone, role),
        )
        conn.commit()
        return cursor.rowcount > 0  # 返回是否插入成功
    finally:
        conn.close()


def delete_user(username):
    """删除用户。"""
    conn = sqlite3.connect(_user_db_path())
    try:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM user WHERE username=?", (username,))
        conn.commit()
        return cursor.rowcount > 0
    finally:
        conn.close()


def update_user(original_username, new_username, new_password, new_phone, new_role):
    """更新用户信息。"""
    conn = sqlite3.connect(_user_db_path())
    try:
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE user SET username=?, password=?, phone=?, role=? WHERE username=?",
            (new_username, new_password, new_phone, new_role, original_username),
        )
        conn.commit()
        return cursor.rowcount > 0
    finally:
        conn.close()


def update_user_phone(username, new_phone):
    """更新用户手机号。"""
    conn = sqlite3.connect(_user_db_path())
    try:
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE user SET phone=? WHERE username=?",
            (new_phone, username),
        )
        conn.commit()
        return cursor.rowcount > 0
    finally:
        conn.close()


# === 借还书相关数据库函数 ===
from datetime import datetime

def get_borrowing_records():
    """获取所有借阅中的记录。"""
    conn = sqlite3.connect(_book_db_path())
    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT record_id, username, bookname, borrow_date, return_date, status
            FROM borrow_records
            WHERE status='借阅中'
            ORDER BY borrow_date DESC
        """)
        return cursor.fetchall()
    finally:
        conn.close()


def get_user_borrow_records(username):
    """获取指定用户的所有借阅历史。"""
    conn = sqlite3.connect(_book_db_path())
    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT bookname, borrow_date, return_date, status
            FROM borrow_records
            WHERE username=?
            ORDER BY borrow_date DESC
        """, (username,))
        return cursor.fetchall()
    finally:
        conn.close()


def update_user_password(username, new_password):
    """更新用户密码。"""
    conn = sqlite3.connect(_user_db_path())
    try:
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE user
            SET password=?
            WHERE username=?
        """, (new_password, username))
        conn.commit()
        return cursor.rowcount > 0
    finally:
        conn.close()


def borrow_book(username, bookname):
    """借书操作：库存减一，插入借阅记录。"""
    conn = sqlite3.connect(_book_db_path())
    try:
        cursor = conn.cursor()
        
        # 检查库存
        cursor.execute("SELECT stock FROM book WHERE bookname=?", (bookname,))
        result = cursor.fetchone()
        if not result:
            return False, "图书不存在"
        
        stock = result[0]
        if stock <= 0:
            return False, "库存不足"
        
        # 库存减一
        cursor.execute("UPDATE book SET stock=stock-1 WHERE bookname=?", (bookname,))
        
        # 更新图书状态（如果库存为0）
        if stock - 1 <= 0:
            cursor.execute("UPDATE book SET status='借出' WHERE bookname=?", (bookname,))
        
        # 插入借阅记录
        borrow_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cursor.execute("""
            INSERT INTO borrow_records (username, bookname, borrow_date, status)
            VALUES (?, ?, ?, '借阅中')
        """, (username, bookname, borrow_date))
        
        conn.commit()
        return True, "借书成功"
    except Exception as e:
        return False, str(e)
    finally:
        conn.close()


def return_book(record_id):
    """还书操作：更新借阅记录状态，库存加一。"""
    conn = sqlite3.connect(_book_db_path())
    try:
        cursor = conn.cursor()
        
        # 获取记录信息
        cursor.execute("SELECT bookname FROM borrow_records WHERE record_id=?", (record_id,))
        result = cursor.fetchone()
        if not result:
            return False, "记录不存在"
        
        bookname = result[0]
        
        # 更新借阅记录
        return_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cursor.execute("""
            UPDATE borrow_records
            SET status='已归还', return_date=?
            WHERE record_id=?
        """, (return_date, record_id))
        
        # 库存加一
        cursor.execute("UPDATE book SET stock=stock+1 WHERE bookname=?", (bookname,))
        
        # 更新图书状态（如果库存>0）
        cursor.execute("SELECT stock FROM book WHERE bookname=?", (bookname,))
        stock = cursor.fetchone()[0]
        if stock > 0:
            cursor.execute("UPDATE book SET status='在馆' WHERE bookname=?", (bookname,))
        
        conn.commit()
        return True, "还书成功"
    except Exception as e:
        return False, str(e)
    finally:
        conn.close()


def _migrate_books_from_user_info():
    """
    若旧版 user_info.db 中遗留有 book 表，把数据迁移到 book_info.db 后删除旧 book 表。
    仅在首次运行时生效，之后是 no-op。
    """
    user_db = _user_db_path()
    if not os.path.exists(user_db):
        return

    src = sqlite3.connect(user_db)
    try:
        src_cur = src.cursor()
        src_cur.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='book'"
        )
        if not src_cur.fetchone():
            return

        rows = src_cur.execute(
            "SELECT bookname, price, author, pubcom FROM book"
        ).fetchall()

        if rows:
            dst = sqlite3.connect(_book_db_path())
            try:
                dst_cur = dst.cursor()
                dst_cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS book (
                        bookname varchar primary key,
                        price varchar,
                        author varchar,
                        pubcom varchar
                    )
                    """
                )
                for row in rows:
                    dst_cur.execute(
                        "INSERT OR REPLACE INTO book "
                        "(bookname, price, author, pubcom) VALUES (?, ?, ?, ?)",
                        row,
                    )
                dst.commit()
            finally:
                dst.close()

        src_cur.execute("DROP TABLE book")
        src.commit()
    except sqlite3.Error:
        pass
    finally:
        src.close()


def init_book_table():
    """
    初始化图书表（连接 book_info.db 并创建 book 表）
    """
    _migrate_books_from_user_info()

    conn = sqlite3.connect(_book_db_path())
    try:
        cursor = conn.cursor()
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS book (
                bookname varchar PRIMARY KEY,
                price varchar,
                author varchar,
                pubcom varchar
            )
            """
        )
        conn.commit()
    finally:
        conn.close()


def migrate_book_fields():
    """
    为 book 表添加 stock（库存，默认1）和 status（状态：在馆/借出）字段
    """
    conn = sqlite3.connect(_book_db_path())
    try:
        cursor = conn.cursor()
        
        # 检查 stock 字段是否存在
        cursor.execute("PRAGMA table_info(book)")
        columns = [col[1] for col in cursor.fetchall()]
        
        if 'stock' not in columns:
            cursor.execute("ALTER TABLE book ADD COLUMN stock INTEGER DEFAULT 1")
            conn.commit()
        
        if 'status' not in columns:
            cursor.execute("ALTER TABLE book ADD COLUMN status TEXT DEFAULT '在馆'")
            conn.commit()
            
    except sqlite3.Error as e:
        print(f"迁移 book 表字段时出错: {e}")
    finally:
        conn.close()


def init_borrow_table():
    """
    初始化借阅记录表 borrow_records
    """
    conn = sqlite3.connect(_book_db_path())
    try:
        cursor = conn.cursor()
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS borrow_records (
                record_id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL,
                bookname TEXT NOT NULL,
                borrow_date TEXT NOT NULL,
                return_date TEXT,
                status TEXT NOT NULL DEFAULT '借阅中'
            )
            """
        )
        conn.commit()
    except sqlite3.Error as e:
        print(f"初始化借阅记录表时出错: {e}")
    finally:
        conn.close()


def init_import_export_request_table():
    """
    初始化导入导出申请表 import_export_requests
    """
    conn = sqlite3.connect(_book_db_path())
    try:
        cursor = conn.cursor()
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS import_export_requests (
                request_id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL,
                request_type TEXT NOT NULL,
                request_date TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT '待处理',
                process_date TEXT,
                remark TEXT
            )
            """
        )
        conn.commit()
    except sqlite3.Error as e:
        print(f"初始化导入导出申请表时出错: {e}")
    finally:
        conn.close()


def init_borrow_request_table():
    """
    初始化借阅申请表 borrow_requests
    """
    conn = sqlite3.connect(_book_db_path())
    try:
        cursor = conn.cursor()
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS borrow_requests (
                request_id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL,
                bookname TEXT NOT NULL,
                request_date TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT '待处理',
                process_date TEXT,
                remark TEXT
            )
            """
        )
        conn.commit()
    except sqlite3.Error as e:
        print(f"初始化借阅申请表时出错: {e}")
    finally:
        conn.close()


def add_borrow_request(username, bookname):
    """
    添加借阅申请
    """
    conn = sqlite3.connect(_book_db_path())
    try:
        cursor = conn.cursor()
        request_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cursor.execute(
            """
            INSERT INTO borrow_requests (username, bookname, request_date, status)
            VALUES (?, ?, ?, '待处理')
            """,
            (username, bookname, request_date),
        )
        conn.commit()
        return True
    except Exception as e:
        print(f"添加借阅申请失败: {e}")
        return False
    finally:
        conn.close()


def get_all_borrow_requests():
    """
    获取所有借阅申请
    """
    conn = sqlite3.connect(_book_db_path())
    try:
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT request_id, username, bookname, request_date, status, process_date, remark
            FROM borrow_requests
            ORDER BY request_date DESC
            """
        )
        return cursor.fetchall()
    finally:
        conn.close()


def get_user_borrow_requests(username):
    """
    获取指定用户的借阅申请
    """
    conn = sqlite3.connect(_book_db_path())
    try:
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT request_id, bookname, request_date, status, process_date, remark
            FROM borrow_requests
            WHERE username=?
            ORDER BY request_date DESC
            """,
            (username,),
        )
        return cursor.fetchall()
    finally:
        conn.close()


def update_borrow_request_status(request_id, status, remark=""):
    """
    更新借阅申请状态
    """
    conn = sqlite3.connect(_book_db_path())
    try:
        cursor = conn.cursor()
        process_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # 如果状态是已批准，则执行借书操作
        if status == "已批准":
            # 获取申请信息
            cursor.execute("SELECT username, bookname FROM borrow_requests WHERE request_id=?", (request_id,))
            result = cursor.fetchone()
            if result:
                username, bookname = result
                # 检查库存
                cursor.execute("SELECT stock FROM book WHERE bookname=?", (bookname,))
                stock_result = cursor.fetchone()
                if not stock_result or stock_result[0] <= 0:
                    return False, "图书不存在或库存不足"
                
                # 执行借书操作
                borrow_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                cursor.execute(
                    "INSERT INTO borrow_records (username, bookname, borrow_date, status) VALUES (?, ?, ?, '借阅中')",
                    (username, bookname, borrow_date)
                )
                cursor.execute("UPDATE book SET stock=stock-1 WHERE bookname=?", (bookname,))
                # 更新图书状态
                cursor.execute("SELECT stock FROM book WHERE bookname=?", (bookname,))
                new_stock = cursor.fetchone()[0]
                if new_stock <= 0:
                    cursor.execute("UPDATE book SET status='借出' WHERE bookname=?", (bookname,))
        
        # 更新申请状态
        cursor.execute(
            """
            UPDATE borrow_requests
            SET status=?, process_date=?, remark=?
            WHERE request_id=?
            """,
            (status, process_date, remark, request_id),
        )
        conn.commit()
        return True, "操作成功"
    except Exception as e:
        print(f"更新借阅申请状态失败: {e}")
        return False, str(e)
    finally:
        conn.close()


def add_import_export_request(username, request_type):
    """
    添加导入导出申请
    """
    conn = sqlite3.connect(_book_db_path())
    try:
        cursor = conn.cursor()
        request_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cursor.execute(
            """
            INSERT INTO import_export_requests (username, request_type, request_date, status)
            VALUES (?, ?, ?, '待处理')
            """,
            (username, request_type, request_date),
        )
        conn.commit()
        return True
    except Exception as e:
        print(f"添加申请失败: {e}")
        return False
    finally:
        conn.close()


def get_all_import_export_requests():
    """
    获取所有导入导出申请
    """
    conn = sqlite3.connect(_book_db_path())
    try:
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT request_id, username, request_type, request_date, status, process_date, remark
            FROM import_export_requests
            ORDER BY request_date DESC
            """
        )
        return cursor.fetchall()
    finally:
        conn.close()


def get_user_import_export_requests(username):
    """
    获取指定用户的导入导出申请
    """
    conn = sqlite3.connect(_book_db_path())
    try:
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT request_id, request_type, request_date, status, process_date, remark
            FROM import_export_requests
            WHERE username=?
            ORDER BY request_date DESC
            """,
            (username,),
        )
        return cursor.fetchall()
    finally:
        conn.close()


def update_import_export_request_status(request_id, status, remark=""):
    """
    更新导入导出申请状态
    """
    conn = sqlite3.connect(_book_db_path())
    try:
        cursor = conn.cursor()
        process_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cursor.execute(
            """
            UPDATE import_export_requests
            SET status=?, process_date=?, remark=?
            WHERE request_id=?
            """,
            (status, process_date, remark, request_id),
        )
        conn.commit()
        return cursor.rowcount > 0
    finally:
        conn.close()


def check_import_export_approved(username, request_type):
    """
    检查用户是否有已批准的对应类型申请
    """
    conn = sqlite3.connect(_book_db_path())
    try:
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT 1
            FROM import_export_requests
            WHERE username=? AND request_type=? AND status='已批准'
            LIMIT 1
            """,
            (username, request_type),
        )
        return cursor.fetchone() is not None
    finally:
        conn.close()


class WelcomeFrame(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        tk.Label(self, text="欢迎页（WelcomeFrame）", font=(None, 16, "bold")).pack(pady=30)


class ListFrame(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        tk.Label(self, text="图书列表页（ListFrame）", font=(None, 16, "bold")).pack(pady=10)

        btn_frame = tk.Frame(self)
        btn_frame.pack(fill=tk.X, padx=12, pady=6)

        self.add_btn = ttk.Button(btn_frame, text="添加图书", command=self.go_add)
        self.add_btn.pack(side=tk.LEFT, padx=4)

        self.edit_btn = ttk.Button(btn_frame, text="修改图书", command=self.edit_book)
        self.edit_btn.pack(side=tk.LEFT, padx=4)

        self.del_btn = ttk.Button(btn_frame, text="删除选中", command=self.delete_books)
        self.del_btn.pack(side=tk.LEFT, padx=4)

        self.refresh_btn = ttk.Button(btn_frame, text="刷新列表", command=self.reload)
        self.refresh_btn.pack(side=tk.LEFT, padx=4)

        self.select_all_btn = ttk.Button(btn_frame, text="全选", command=self.select_all)
        self.select_all_btn.pack(side=tk.LEFT, padx=4)

        self.deselect_all_btn = ttk.Button(btn_frame, text="取消全选", command=self.deselect_all)
        self.deselect_all_btn.pack(side=tk.LEFT, padx=4)

        columns = ("select", "bookname", "author", "price", "pubcom", "stock", "status")
        self.tree = ttk.Treeview(self, columns=columns, show="headings", height=18)
        self.tree.heading("select", text="选择")
        self.tree.heading("bookname", text="书名")
        self.tree.heading("author", text="作者")
        self.tree.heading("price", text="价格")
        self.tree.heading("pubcom", text="出版社")
        self.tree.heading("stock", text="库存")
        self.tree.heading("status", text="状态")
        self.tree.column("select", width=50, anchor=tk.CENTER)
        self.tree.column("bookname", width=120, anchor=tk.CENTER)
        self.tree.column("author", width=100, anchor=tk.CENTER)
        self.tree.column("price", width=70, anchor=tk.CENTER)
        self.tree.column("pubcom", width=120, anchor=tk.CENTER)
        self.tree.column("stock", width=50, anchor=tk.CENTER)
        self.tree.column("status", width=60, anchor=tk.CENTER)

        # 滚动条
        scroll = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scroll.set)

        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=12, pady=8)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # 存储选中状态
        self.selected_items = set()

        # 绑定点击事件切换选中
        self.tree.bind("<Button-1>", self.on_click)

        self.reload()
        if self.master.role != "admin":
            self.add_btn.config(state=tk.DISABLED)
            self.edit_btn.config(state=tk.DISABLED)
            self.del_btn.config(state=tk.DISABLED)
            self.select_all_btn.config(state=tk.DISABLED)
            self.deselect_all_btn.config(state=tk.DISABLED)

    def on_click(self, event):
        """点击事件 - 切换选中状态"""
        item = self.tree.identify_row(event.y)
        if not item:
            return

        # 切换选中状态
        if item in self.selected_items:
            self.selected_items.remove(item)
            self.tree.set(item, "select", "")
        else:
            self.selected_items.add(item)
            self.tree.set(item, "select", "☑")

    def select_all(self):
        """全选"""
        self.selected_items.clear()
        for item in self.tree.get_children():
            self.selected_items.add(item)
            self.tree.set(item, "select", "☑")

    def deselect_all(self):
        """取消全选"""
        for item in self.tree.get_children():
            self.tree.set(item, "select", "")
        self.selected_items.clear()

    def _db_path(self):
        return _book_db_path()

    def go_add(self):
        self.master.showFrame(self.master.add_frame)

    def reload(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.selected_items.clear()

        conn = sqlite3.connect(self._db_path())
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT bookname, author, price, pubcom, stock, status FROM book")
            for row in cursor.fetchall():
                bookname, author, price, pubcom, stock, status = row
                self.tree.insert("", tk.END, values=("", bookname, author, price, pubcom, stock, status))
        finally:
            conn.close()

    def edit_book(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要修改的图书")
            return

        # 获取选中的图书信息
        item = selected[0]
        bookname, author, price, pubcom, stock, status = self.tree.item(item, "values")[1:]
        original_bookname = bookname

        # 创建修改对话框
        dialog = tk.Toplevel(self)
        dialog.title("修改图书")
        dialog.geometry("450x400")
        dialog.resizable(False, False)
        dialog.transient(self)
        dialog.grab_set()

        form = tk.Frame(dialog)
        form.pack(pady=20)

        tk.Label(form, text="书名").grid(row=0, column=0, padx=8, pady=8, sticky=tk.E)
        bookname_entry = ttk.Entry(form, width=28)
        bookname_entry.grid(row=0, column=1, padx=8, pady=8)
        bookname_entry.insert(0, str(bookname))

        tk.Label(form, text="作者").grid(row=1, column=0, padx=8, pady=8, sticky=tk.E)
        author_entry = ttk.Entry(form, width=28)
        author_entry.grid(row=1, column=1, padx=8, pady=8)
        author_entry.insert(0, str(author))

        tk.Label(form, text="价格").grid(row=2, column=0, padx=8, pady=8, sticky=tk.E)
        price_entry = ttk.Entry(form, width=28)
        price_entry.grid(row=2, column=1, padx=8, pady=8)
        price_entry.insert(0, str(price))

        tk.Label(form, text="出版社").grid(row=3, column=0, padx=8, pady=8, sticky=tk.E)
        pubcom_entry = ttk.Entry(form, width=28)
        pubcom_entry.grid(row=3, column=1, padx=8, pady=8)
        pubcom_entry.insert(0, str(pubcom))

        tk.Label(form, text="库存").grid(row=4, column=0, padx=8, pady=8, sticky=tk.E)
        stock_entry = ttk.Entry(form, width=28)
        stock_entry.grid(row=4, column=1, padx=8, pady=8)
        stock_entry.insert(0, str(stock))

        tk.Label(form, text="状态").grid(row=5, column=0, padx=8, pady=8, sticky=tk.E)
        status_var = tk.StringVar(value=str(status))
        status_combo = ttk.Combobox(form, textvariable=status_var, values=["在馆", "借出"], width=25, state="readonly")
        status_combo.grid(row=5, column=1, padx=8, pady=8)

        # 确保对话框显示后再设置焦点
        dialog.update()
        bookname_entry.focus_force()

        def save():
            new_bookname = bookname_entry.get().strip()
            new_author = author_entry.get().strip()
            new_price = price_entry.get().strip()
            new_pubcom = pubcom_entry.get().strip()
            new_stock = stock_entry.get().strip()
            new_status = status_var.get().strip()

            if not all([new_bookname, new_author, new_price, new_pubcom, new_stock]):
                messagebox.showwarning("提示", "请完整填写图书信息")
                return

            try:
                new_stock = int(new_stock)
            except ValueError:
                messagebox.showwarning("提示", "库存必须是整数")
                return

            conn = sqlite3.connect(self._db_path())
            try:
                cursor = conn.cursor()
                cursor.execute(
                    "UPDATE book SET bookname=?, price=?, author=?, pubcom=?, stock=?, status=? WHERE bookname=?",
                    (new_bookname, new_price, new_author, new_pubcom, new_stock, new_status, original_bookname)
                )
                conn.commit()
            finally:
                conn.close()

            messagebox.showinfo("成功", "图书修改成功！")
            dialog.destroy()
            self.reload()

        btn_frame = tk.Frame(dialog)
        btn_frame.pack(pady=20)
        ttk.Button(btn_frame, text="保存", command=save).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="取消", command=dialog.destroy).pack(side=tk.LEFT, padx=10)

    def delete_books(self):
        """批量删除图书"""
        if not self.selected_items:
            messagebox.showwarning("提示", "请先选择要删除的图书")
            return

        # 获取选中的图书
        selected_books = []
        for item in self.selected_items:
            bookname = self.tree.item(item, "values")[1]
            selected_books.append(bookname)

        # 确认对话框
        book_list_text = "\n".join([f"  • {name}" for name in selected_books[:10]])
        if len(selected_books) > 10:
            book_list_text += f"\n  ... 还有 {len(selected_books) - 10} 本"

        confirm_msg = (
            f"确定要删除以下 {len(selected_books)} 本图书吗？\n\n"
            f"{book_list_text}\n\n"
            f"此操作不可恢复！"
        )

        if not messagebox.askyesno("确认删除", confirm_msg):
            return

        # 删除图书
        conn = sqlite3.connect(self._db_path())
        try:
            cursor = conn.cursor()
            success_count = 0
            fail_count = 0

            for bookname in selected_books:
                cursor.execute("DELETE FROM book WHERE bookname=?", (bookname,))
                if cursor.rowcount > 0:
                    success_count += 1
                else:
                    fail_count += 1

            conn.commit()

            messagebox.showinfo("结果", f"删除完成！成功 {success_count} 本，失败 {fail_count} 本")
        except Exception as e:
            messagebox.showerror("错误", f"删除失败: {e}")
        finally:
            conn.close()

        self.reload()


class AddFrame(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        tk.Label(self, text="图书添加页（AddFrame）", font=(None, 16, "bold")).pack(pady=10)

        form = tk.Frame(self)
        form.pack(pady=20)

        tk.Label(form, text="书名").grid(row=0, column=0, padx=8, pady=8, sticky=tk.E)
        self.bookname_entry = ttk.Entry(form, width=28)
        self.bookname_entry.grid(row=0, column=1, padx=8, pady=8)
        self.set_placeholder(self.bookname_entry, '请输入书名')

        tk.Label(form, text="作者").grid(row=1, column=0, padx=8, pady=8, sticky=tk.E)
        self.author_entry = ttk.Entry(form, width=28)
        self.author_entry.grid(row=1, column=1, padx=8, pady=8)
        self.set_placeholder(self.author_entry, '请输入作者')

        tk.Label(form, text="价格").grid(row=2, column=0, padx=8, pady=8, sticky=tk.E)
        self.price_entry = ttk.Entry(form, width=28)
        self.price_entry.grid(row=2, column=1, padx=8, pady=8)
        self.set_placeholder(self.price_entry, '请输入价格')

        tk.Label(form, text="出版社").grid(row=3, column=0, padx=8, pady=8, sticky=tk.E)
        self.pubcom_entry = ttk.Entry(form, width=28)
        self.pubcom_entry.grid(row=3, column=1, padx=8, pady=8)
        self.set_placeholder(self.pubcom_entry, '请输入出版社')

        tk.Label(form, text="库存").grid(row=4, column=0, padx=8, pady=8, sticky=tk.E)
        self.stock_entry = ttk.Entry(form, width=28)
        self.stock_entry.grid(row=4, column=1, padx=8, pady=8)
        self.stock_entry.insert(0, '1')

        tk.Label(form, text="状态").grid(row=5, column=0, padx=8, pady=8, sticky=tk.E)
        self.status_var = tk.StringVar(value="在馆")
        self.status_combo = ttk.Combobox(form, textvariable=self.status_var, values=["在馆", "借出"], width=25, state="readonly")
        self.status_combo.grid(row=5, column=1, padx=8, pady=8)

        action = tk.Frame(self)
        action.pack(pady=6)
        ttk.Button(action, text="保存图书", command=self.save_book).pack(side=tk.LEFT, padx=6)
        ttk.Button(
            action,
            text="返回列表",
            command=lambda: self.master.showFrame(self.master.list_frame)
        ).pack(side=tk.LEFT, padx=6)

    def set_placeholder(self, entry, text):
        """为Entry组件添加占位符"""
        entry.insert(0, text)

        def on_focus_in(event):
            if entry.get() == text:
                entry.delete(0, tk.END)

        def on_focus_out(event):
            if not entry.get():
                entry.insert(0, text)

        entry.bind("<FocusIn>", on_focus_in, add='+')
        entry.bind("<FocusOut>", on_focus_out, add='+')

    def _db_path(self):
        return _book_db_path()

    def save_book(self):
        # 获取Entry真实内容，若等于占位符内容则视为空
        bookname = self.bookname_entry.get().strip()
        bookname = '' if bookname == '请输入书名' else bookname
        
        author = self.author_entry.get().strip()
        author = '' if author == '请输入作者' else author
        
        price = self.price_entry.get().strip()
        price = '' if price == '请输入价格' else price
        
        pubcom = self.pubcom_entry.get().strip()
        pubcom = '' if pubcom == '请输入出版社' else pubcom
        
        stock = self.stock_entry.get().strip()
        status = self.status_var.get().strip()

        if not bookname:
            messagebox.showwarning("提示", "请输入书名")
            self.bookname_entry.focus_set()
            return
        if not author:
            messagebox.showwarning("提示", "请输入作者")
            self.author_entry.focus_set()
            return
        if not price:
            messagebox.showwarning("提示", "请输入价格")
            self.price_entry.focus_set()
            return
        if not pubcom:
            messagebox.showwarning("提示", "请输入出版社")
            self.pubcom_entry.focus_set()
            return
        if not stock:
            messagebox.showwarning("提示", "请输入库存")
            self.stock_entry.focus_set()
            return

        try:
            stock = int(stock)
        except ValueError:
            messagebox.showwarning("提示", "库存必须是整数")
            self.stock_entry.focus_set()
            return

        conn = sqlite3.connect(self._db_path())
        try:
            cursor = conn.cursor()
            cursor.execute(
                "INSERT OR REPLACE INTO book (bookname, price, author, pubcom, stock, status) VALUES (?, ?, ?, ?, ?, ?)",
                (bookname, price, author, pubcom, stock, status),
            )
            conn.commit()
        finally:
            conn.close()

        messagebox.showinfo("成功", "图书添加成功")
        # 清空并重置输入框
        self.bookname_entry.delete(0, tk.END)
        self.set_placeholder(self.bookname_entry, '请输入书名')
        
        self.author_entry.delete(0, tk.END)
        self.set_placeholder(self.author_entry, '请输入作者')
        
        self.price_entry.delete(0, tk.END)
        self.set_placeholder(self.price_entry, '请输入价格')
        
        self.pubcom_entry.delete(0, tk.END)
        self.set_placeholder(self.pubcom_entry, '请输入出版社')
        
        self.stock_entry.delete(0, tk.END)
        self.stock_entry.insert(0, '1')
        
        self.status_var.set("在馆")
        
        self.master.list_frame.reload()
        self.master.showFrame(self.master.list_frame)


class DataFrame(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        tk.Label(self, text="数据分析页（DataFrame）", font=(None, 16, "bold")).pack(pady=10)
        
        # 按钮区域
        btn_frame = tk.Frame(self)
        btn_frame.pack(pady=10)
        
        self.pub_pie_btn = ttk.Button(
            btn_frame,
            text="出版社发行数量统计",
            command=self.show_pie_chart
        )
        self.pub_pie_btn.pack(side=tk.LEFT, padx=5)
        
        self.save_chart_btn = ttk.Button(
            btn_frame,
            text="保存图表",
            command=self.save_chart,
            state=tk.DISABLED
        )
        self.save_chart_btn.pack(side=tk.LEFT, padx=5)
        
        # 图表显示区域 - 使用Frame容器
        self.chart_container = tk.Frame(self)
        self.chart_container.pack(pady=10, fill=tk.BOTH, expand=True)
        
        self.current_figure = None
        self.current_canvas = None
        self.current_chart_path = None

    def show_pie_chart(self):
        conn = sqlite3.connect(_book_db_path())
        try:
            df = pd.read_sql_query("SELECT pubcom FROM book", conn)
        finally:
            conn.close()

        pub_series = df["pubcom"].dropna().astype(str).str.strip()
        pub_series = pub_series[pub_series != ""]
        counts = pub_series.value_counts()

        if counts.empty:
            messagebox.showinfo("提示", "当前没有可用于统计的出版社数据")
            return

        # 清除之前的图表
        if self.current_canvas:
            self.current_canvas.get_tk_widget().pack_forget()
            self.current_canvas = None
        if self.current_figure:
            plt.close(self.current_figure)
            self.current_figure = None

        # 创建新图表
        self.current_figure = plt.figure(figsize=(8, 6))
        plt.pie(
            counts.values,
            labels=counts.index,
            autopct="%.1f%%"
        )
        plt.title("各出版社图书数量占比")
        
        # 保存图表到项目内的images文件夹
        images_dir = os.path.join(_project_root(), "images")
        os.makedirs(images_dir, exist_ok=True)
        self.current_chart_path = os.path.join(images_dir, "pub_pie_chart.png")
        plt.savefig(self.current_chart_path, dpi=100, bbox_inches='tight')
        
        # 直接在tkinter中显示图表
        self.current_canvas = FigureCanvasTkAgg(self.current_figure, master=self.chart_container)
        self.current_canvas.draw()
        self.current_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        
        self.save_chart_btn.config(state=tk.NORMAL)
    
    def save_chart(self):
        """另存为图表"""
        if not self.current_chart_path or not os.path.exists(self.current_chart_path):
            messagebox.showwarning("提示", "请先生成图表")
            return
        
        file_path = asksaveasfilename(
            title="保存图表",
            defaultextension=".png",
            filetypes=[("PNG 图片", "*.png"), ("JPEG 图片", "*.jpg")],
        )
        
        if file_path:
            import shutil
            shutil.copy2(self.current_chart_path, file_path)
            messagebox.showinfo("成功", "图表保存成功！")


class StudentBookListFrame(tk.Frame):
    """学生图书列表界面 - 可以申请借阅"""
    def __init__(self, master=None):
        super().__init__(master)
        tk.Label(self, text="图书列表", font=(None, 16, "bold")).pack(pady=10)
        
        # 按钮区域
        btn_frame = tk.Frame(self)
        btn_frame.pack(fill=tk.X, padx=12, pady=6)
        
        self.refresh_btn = ttk.Button(btn_frame, text="刷新列表", command=self.reload)
        self.refresh_btn.pack(side=tk.LEFT, padx=4)
        
        self.apply_borrow_btn = ttk.Button(btn_frame, text="申请借阅", command=self.apply_borrow)
        self.apply_borrow_btn.pack(side=tk.LEFT, padx=4)
        
        self.select_all_btn = ttk.Button(btn_frame, text="全选", command=self.select_all)
        self.select_all_btn.pack(side=tk.LEFT, padx=4)
        
        self.deselect_all_btn = ttk.Button(btn_frame, text="取消全选", command=self.deselect_all)
        self.deselect_all_btn.pack(side=tk.LEFT, padx=4)
        
        # 表格区域
        columns = ("select", "bookname", "author", "price", "pubcom", "stock", "status")
        self.tree = ttk.Treeview(self, columns=columns, show="headings", height=18)
        self.tree.heading("select", text="选择")
        self.tree.heading("bookname", text="书名")
        self.tree.heading("author", text="作者")
        self.tree.heading("price", text="价格")
        self.tree.heading("pubcom", text="出版社")
        self.tree.heading("stock", text="库存")
        self.tree.heading("status", text="状态")
        self.tree.column("select", width=50, anchor=tk.CENTER)
        self.tree.column("bookname", width=150, anchor=tk.W)
        self.tree.column("author", width=100, anchor=tk.W)
        self.tree.column("price", width=70, anchor=tk.CENTER)
        self.tree.column("pubcom", width=120, anchor=tk.W)
        self.tree.column("stock", width=60, anchor=tk.CENTER)
        self.tree.column("status", width=70, anchor=tk.CENTER)
        
        # 滚动条
        scroll = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scroll.set)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=12, pady=8)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 存储选中状态
        self.selected_items = set()
        
        # 绑定点击事件切换选中
        self.tree.bind("<Button-1>", self.on_click)
        
        self.reload()
    
    def on_click(self, event):
        """点击事件 - 切换选中状态"""
        item = self.tree.identify_row(event.y)
        if not item:
            return
        
        # 切换选中状态
        if item in self.selected_items:
            self.selected_items.remove(item)
            self.tree.set(item, "select", "")
        else:
            self.selected_items.add(item)
            self.tree.set(item, "select", "☑")
    
    def select_all(self):
        """全选"""
        self.selected_items.clear()
        for item in self.tree.get_children():
            self.selected_items.add(item)
            self.tree.set(item, "select", "☑")
    
    def deselect_all(self):
        """取消全选"""
        for item in self.tree.get_children():
            self.tree.set(item, "select", "")
        self.selected_items.clear()
    
    def reload(self):
        """刷新图书列表"""
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.selected_items.clear()
        
        conn = sqlite3.connect(_book_db_path())
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT bookname, author, price, pubcom, stock, status FROM book")
            for row in cursor.fetchall():
                bookname, author, price, pubcom, stock, status = row
                self.tree.insert("", tk.END, values=("", bookname, author, price, pubcom, stock, status))
        finally:
            conn.close()
    
    def apply_borrow(self):
        """申请借阅"""
        if not self.selected_items:
            messagebox.showwarning("提示", "请先选择要借阅的图书")
            return
        
        # 获取选中的图书
        selected_books = []
        for item in self.selected_items:
            bookname = self.tree.item(item, "values")[1]
            status = self.tree.item(item, "values")[6]
            stock = int(self.tree.item(item, "values")[5])
            
            if status == "借出" or stock <= 0:
                messagebox.showwarning("提示", f"图书《{bookname}》库存不足或已借出，无法申请")
                return
            
            selected_books.append(bookname)
        
        if not selected_books:
            return
        
        # 确认对话框
        book_list_text = "\n".join([f"  • {name}" for name in selected_books[:10]])
        if len(selected_books) > 10:
            book_list_text += f"\n  ... 还有 {len(selected_books) - 10} 本"
        
        confirm_msg = (
            f"确定要申请借阅以下 {len(selected_books)} 本图书吗？\n\n"
            f"{book_list_text}\n\n"
            f"请等待管理员审批。"
        )
        
        if not messagebox.askyesno("确认申请", confirm_msg):
            return
        
        # 提交申请
        success_count = 0
        fail_count = 0
        for bookname in selected_books:
            if add_borrow_request(self.master.current_user, bookname):
                success_count += 1
            else:
                fail_count += 1
        
        if success_count > 0:
            messagebox.showinfo("成功", f"申请提交成功！成功 {success_count} 本，失败 {fail_count} 本")
        else:
            messagebox.showerror("错误", "申请提交失败！")
        
        self.deselect_all()


class MyBorrowRequestsFrame(tk.Frame):
    """我的借阅申请界面"""
    def __init__(self, master=None):
        super().__init__(master)
        tk.Label(self, text="我的借阅申请", font=(None, 16, "bold")).pack(pady=10)
        
        # 按钮区域
        btn_frame = tk.Frame(self)
        btn_frame.pack(fill=tk.X, padx=12, pady=6)
        
        self.refresh_btn = ttk.Button(btn_frame, text="刷新列表", command=self.reload)
        self.refresh_btn.pack(side=tk.LEFT, padx=4)
        
        # 表格区域
        columns = ("request_id", "bookname", "request_date", "status", "process_date", "remark")
        self.tree = ttk.Treeview(self, columns=columns, show="headings", height=18)
        self.tree.heading("request_id", text="申请ID")
        self.tree.heading("bookname", text="图书名称")
        self.tree.heading("request_date", text="申请时间")
        self.tree.heading("status", text="状态")
        self.tree.heading("process_date", text="处理时间")
        self.tree.heading("remark", text="备注")
        self.tree.column("request_id", width=60, anchor=tk.CENTER)
        self.tree.column("bookname", width=200, anchor=tk.W)
        self.tree.column("request_date", width=140, anchor=tk.CENTER)
        self.tree.column("status", width=80, anchor=tk.CENTER)
        self.tree.column("process_date", width=140, anchor=tk.CENTER)
        self.tree.column("remark", width=200, anchor=tk.W)
        
        # 滚动条
        scroll = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scroll.set)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=12, pady=8)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.reload()
    
    def reload(self):
        """刷新申请列表"""
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        requests = get_user_borrow_requests(self.master.current_user)
        for req in requests:
            request_id, bookname, request_date, status, process_date, remark = req
            process_date_display = process_date if process_date else "-"
            remark_display = remark if remark else "-"
            self.tree.insert("", tk.END, values=(request_id, bookname, request_date, status, process_date_display, remark_display))


class BorrowRequestApprovalFrame(tk.Frame):
    """借阅申请审批界面（仅管理员）"""
    def __init__(self, master=None):
        super().__init__(master)
        tk.Label(self, text="借阅申请审批", font=(None, 16, "bold")).pack(pady=10)
        
        # 按钮区域
        btn_frame = tk.Frame(self)
        btn_frame.pack(fill=tk.X, padx=12, pady=6)
        
        self.approve_btn = ttk.Button(btn_frame, text="批准申请", command=self.approve_requests)
        self.approve_btn.pack(side=tk.LEFT, padx=4)
        
        self.reject_btn = ttk.Button(btn_frame, text="拒绝申请", command=self.reject_requests)
        self.reject_btn.pack(side=tk.LEFT, padx=4)
        
        self.refresh_btn = ttk.Button(btn_frame, text="刷新列表", command=self.reload)
        self.refresh_btn.pack(side=tk.LEFT, padx=4)
        
        self.select_all_btn = ttk.Button(btn_frame, text="全选", command=self.select_all)
        self.select_all_btn.pack(side=tk.LEFT, padx=4)
        
        self.deselect_all_btn = ttk.Button(btn_frame, text="取消全选", command=self.deselect_all)
        self.deselect_all_btn.pack(side=tk.LEFT, padx=4)
        
        # 表格区域
        columns = ("select", "request_id", "username", "bookname", "request_date", "status", "process_date", "remark")
        self.tree = ttk.Treeview(self, columns=columns, show="headings", height=18)
        self.tree.heading("select", text="选择")
        self.tree.heading("request_id", text="申请ID")
        self.tree.heading("username", text="申请人")
        self.tree.heading("bookname", text="图书名称")
        self.tree.heading("request_date", text="申请时间")
        self.tree.heading("status", text="状态")
        self.tree.heading("process_date", text="处理时间")
        self.tree.heading("remark", text="备注")
        self.tree.column("select", width=50, anchor=tk.CENTER)
        self.tree.column("request_id", width=60, anchor=tk.CENTER)
        self.tree.column("username", width=100, anchor=tk.W)
        self.tree.column("bookname", width=180, anchor=tk.W)
        self.tree.column("request_date", width=140, anchor=tk.CENTER)
        self.tree.column("status", width=80, anchor=tk.CENTER)
        self.tree.column("process_date", width=140, anchor=tk.CENTER)
        self.tree.column("remark", width=150, anchor=tk.W)
        
        # 滚动条
        scroll = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scroll.set)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=12, pady=8)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 存储选中状态
        self.selected_items = set()
        
        # 绑定点击事件切换选中
        self.tree.bind("<Button-1>", self.on_click)
        
        self.reload()
    
    def on_click(self, event):
        """点击事件 - 切换选中状态"""
        item = self.tree.identify_row(event.y)
        if not item:
            return
        
        # 切换选中状态
        if item in self.selected_items:
            self.selected_items.remove(item)
            self.tree.set(item, "select", "")
        else:
            self.selected_items.add(item)
            self.tree.set(item, "select", "☑")
    
    def select_all(self):
        """全选"""
        self.selected_items.clear()
        for item in self.tree.get_children():
            status = self.tree.item(item, "values")[5]
            if status == "待处理":
                self.selected_items.add(item)
                self.tree.set(item, "select", "☑")
    
    def deselect_all(self):
        """取消全选"""
        for item in self.tree.get_children():
            self.tree.set(item, "select", "")
        self.selected_items.clear()
    
    def reload(self):
        """刷新申请列表"""
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.selected_items.clear()
        
        requests = get_all_borrow_requests()
        for req in requests:
            request_id, username, bookname, request_date, status, process_date, remark = req
            process_date_display = process_date if process_date else "-"
            remark_display = remark if remark else "-"
            self.tree.insert("", tk.END, values=("", request_id, username, bookname, request_date, status, process_date_display, remark_display))
    
    def approve_requests(self):
        """批量批准申请"""
        if not self.selected_items:
            messagebox.showwarning("提示", "请先选择要批准的申请")
            return
        
        # 获取选中的申请
        selected_requests = []
        for item in self.selected_items:
            request_id = self.tree.item(item, "values")[1]
            status = self.tree.item(item, "values")[5]
            
            if status != "待处理":
                messagebox.showwarning("提示", f"申请 {request_id} 已处理，无法再次处理")
                return
            
            selected_requests.append(request_id)
        
        if not selected_requests:
            return
        
        # 确认对话框
        confirm_msg = f"确定要批准以下 {len(selected_requests)} 个借阅申请吗？"
        if not messagebox.askyesno("确认批准", confirm_msg):
            return
        
        # 获取备注（可选）
        from tkinter.simpledialog import askstring
        remark = askstring("备注", "请输入备注（可选）：")
        if remark is None:
            remark = ""
        
        # 处理申请
        success_count = 0
        fail_count = 0
        fail_messages = []
        
        conn = sqlite3.connect(_book_db_path())
        try:
            cursor = conn.cursor()
            
            for request_id in selected_requests:
                process_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                
                # 获取申请信息
                cursor.execute("SELECT username, bookname FROM borrow_requests WHERE request_id=?", (request_id,))
                result = cursor.fetchone()
                if not result:
                    fail_count += 1
                    fail_messages.append(f"申请 {request_id}: 申请不存在")
                    continue
                
                username, bookname = result
                
                # 检查库存
                cursor.execute("SELECT stock FROM book WHERE bookname=?", (bookname,))
                stock_result = cursor.fetchone()
                if not stock_result or stock_result[0] <= 0:
                    fail_count += 1
                    fail_messages.append(f"申请 {request_id}: 图书《{bookname}》库存不足")
                    continue
                
                # 执行借书操作
                borrow_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                cursor.execute(
                    "INSERT INTO borrow_records (username, bookname, borrow_date, status) VALUES (?, ?, ?, '借阅中')",
                    (username, bookname, borrow_date)
                )
                cursor.execute("UPDATE book SET stock=stock-1 WHERE bookname=?", (bookname,))
                
                # 更新图书状态
                cursor.execute("SELECT stock FROM book WHERE bookname=?", (bookname,))
                new_stock = cursor.fetchone()[0]
                if new_stock <= 0:
                    cursor.execute("UPDATE book SET status='借出' WHERE bookname=?", (bookname,))
                
                # 更新申请状态
                cursor.execute(
                    "UPDATE borrow_requests SET status=?, process_date=?, remark=? WHERE request_id=?",
                    ("已批准", process_date, remark, request_id)
                )
                
                success_count += 1
            
            conn.commit()
        except Exception as e:
            conn.rollback()
            fail_count = len(selected_requests)
            fail_messages.append(f"处理失败: {str(e)}")
        finally:
            conn.close()
        
        # 显示结果
        result_msg = f"处理完成！\n成功: {success_count} 个\n失败: {fail_count} 个"
        if fail_messages:
            result_msg += "\n\n失败详情:\n" + "\n".join(fail_messages[:10])
            if len(fail_messages) > 10:
                result_msg += f"\n... 还有 {len(fail_messages) - 10} 条"
        
        messagebox.showinfo("处理结果", result_msg)
        self.reload()
    
    def reject_requests(self):
        """批量拒绝申请"""
        if not self.selected_items:
            messagebox.showwarning("提示", "请先选择要拒绝的申请")
            return
        
        # 获取选中的申请
        selected_requests = []
        for item in self.selected_items:
            request_id = self.tree.item(item, "values")[1]
            status = self.tree.item(item, "values")[5]
            
            if status != "待处理":
                messagebox.showwarning("提示", f"申请 {request_id} 已处理，无法再次处理")
                return
            
            selected_requests.append(request_id)
        
        if not selected_requests:
            return
        
        # 确认对话框
        confirm_msg = f"确定要拒绝以下 {len(selected_requests)} 个借阅申请吗？"
        if not messagebox.askyesno("确认拒绝", confirm_msg):
            return
        
        # 获取备注（可选）
        from tkinter.simpledialog import askstring
        remark = askstring("备注", "请输入备注（可选）：")
        if remark is None:
            remark = ""
        
        # 处理申请
        success_count = 0
        fail_count = 0
        
        process_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        for request_id in selected_requests:
            result, message = update_borrow_request_status(request_id, "已拒绝", remark)
            if result:
                success_count += 1
            else:
                fail_count += 1
        
        messagebox.showinfo("处理结果", f"处理完成！\n成功: {success_count} 个\n失败: {fail_count} 个")
        self.reload()


class AboutFrame(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        tk.Label(self, text="关于我们（AboutFrame）", font=(None, 16, "bold")).pack(pady=30)


class UserManageFrame(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        tk.Label(self, text="用户管理（UserManageFrame）", font=(None, 16, "bold")).pack(pady=10)

        # 上半部分：操作区
        btn_frame = tk.Frame(self)
        btn_frame.pack(fill=tk.X, padx=12, pady=10)

        self.add_btn = ttk.Button(btn_frame, text="添加用户", command=self.add_user_dialog)
        self.add_btn.pack(side=tk.LEFT, padx=4)

        self.edit_btn = ttk.Button(btn_frame, text="修改用户", command=self.edit_user_dialog)
        self.edit_btn.pack(side=tk.LEFT, padx=4)

        self.delete_btn = ttk.Button(btn_frame, text="删除选中用户", command=self.delete_user)
        self.delete_btn.pack(side=tk.LEFT, padx=4)

        self.refresh_btn = ttk.Button(btn_frame, text="刷新列表", command=self.reload)
        self.refresh_btn.pack(side=tk.LEFT, padx=4)

        # 下半部分：Treeview展示区
        columns = ("username", "phone", "role")
        self.tree = ttk.Treeview(self, columns=columns, show="headings", height=18)
        self.tree.heading("username", text="用户名")
        self.tree.heading("phone", text="手机号")
        self.tree.heading("role", text="角色")
        self.tree.column("username", width=150, anchor=tk.CENTER)
        self.tree.column("phone", width=150, anchor=tk.CENTER)
        self.tree.column("role", width=100, anchor=tk.CENTER)
        self.tree.pack(fill=tk.BOTH, expand=True, padx=12, pady=8)

        self.reload()

    def reload(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        users = load_all_users()
        for user in users:
            username, _, phone, role = user
            role_display = "学生" if role == "student" else "管理员"
            self.tree.insert("", tk.END, values=(username, phone, role_display))

    def add_user_dialog(self):
        # 创建添加用户的弹窗
        dialog = tk.Toplevel(self)
        dialog.title("添加用户")
        dialog.geometry("400x350")
        dialog.resizable(False, False)
        dialog.transient(self)
        dialog.grab_set()

        tk.Label(dialog, text="用户名", font=(None, 10)).pack(pady=(20, 5))
        username_entry = ttk.Entry(dialog, width=30)
        username_entry.pack(pady=5)

        tk.Label(dialog, text="手机号", font=(None, 10)).pack(pady=5)
        phone_entry = ttk.Entry(dialog, width=30)
        phone_entry.pack(pady=5)

        tk.Label(dialog, text="密码", font=(None, 10)).pack(pady=5)
        password_entry = ttk.Entry(dialog, width=30, show="*")
        password_entry.insert(0, "123456")
        password_entry.pack(pady=5)

        tk.Label(dialog, text="角色", font=(None, 10)).pack(pady=5)
        role_var = tk.StringVar(value="student")
        role_combo = ttk.Combobox(dialog, textvariable=role_var, values=["student", "admin"], width=27, state="readonly")
        role_combo.pack(pady=5)

        # 确保对话框显示后再设置焦点
        dialog.update()
        username_entry.focus_force()

        def save():
            username = username_entry.get().strip()
            phone = phone_entry.get().strip()
            password = password_entry.get().strip() or "123456"
            role = role_var.get().strip()

            if not username:
                messagebox.showwarning("提示", "请输入用户名")
                username_entry.focus_force()
                return
            if not phone:
                messagebox.showwarning("提示", "请输入手机号")
                phone_entry.focus_force()
                return

            if add_user(username, password, phone, role):
                messagebox.showinfo("成功", "用户添加成功！")
                dialog.destroy()
                self.reload()
            else:
                messagebox.showerror("错误", "用户名已存在！")

        btn_frame = tk.Frame(dialog)
        btn_frame.pack(pady=20)
        ttk.Button(btn_frame, text="确定", command=save).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="取消", command=dialog.destroy).pack(side=tk.LEFT, padx=10)

    def edit_user_dialog(self):
        # 获取选中的用户
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要修改的用户")
            return

        username = self.tree.item(selected[0], "values")[0]
        # 获取用户完整信息
        users = load_all_users()
        user_info = None
        for user in users:
            if user[0] == username:
                user_info = user
                break

        if not user_info:
            messagebox.showerror("错误", "用户信息不存在")
            return

        original_username, original_password, original_phone, original_role = user_info

        # 创建修改用户的弹窗
        dialog = tk.Toplevel(self)
        dialog.title("修改用户")
        dialog.geometry("400x380")
        dialog.resizable(False, False)
        dialog.transient(self)
        dialog.grab_set()

        tk.Label(dialog, text="用户名", font=(None, 10)).pack(pady=(20, 5))
        username_entry = ttk.Entry(dialog, width=30)
        username_entry.pack(pady=5)
        username_entry.insert(0, original_username)

        tk.Label(dialog, text="手机号", font=(None, 10)).pack(pady=5)
        phone_entry = ttk.Entry(dialog, width=30)
        phone_entry.pack(pady=5)
        phone_entry.insert(0, original_phone)

        tk.Label(dialog, text="密码（留空则不修改）", font=(None, 10)).pack(pady=5)
        password_entry = ttk.Entry(dialog, width=30, show="*")
        password_entry.pack(pady=5)

        tk.Label(dialog, text="角色", font=(None, 10)).pack(pady=5)
        role_var = tk.StringVar(value=original_role)
        role_combo = ttk.Combobox(dialog, textvariable=role_var, values=["student", "admin"], width=27, state="readonly")
        role_combo.pack(pady=5)

        # 确保对话框显示后再设置焦点
        dialog.update()
        username_entry.focus_force()

        def save():
            new_username = username_entry.get().strip()
            new_phone = phone_entry.get().strip()
            new_password = password_entry.get().strip()
            new_role = role_var.get().strip()

            if not new_username:
                messagebox.showwarning("提示", "请输入用户名")
                username_entry.focus_force()
                return
            if not new_phone:
                messagebox.showwarning("提示", "请输入手机号")
                phone_entry.focus_force()
                return

            # 如果密码留空，则使用原密码
            if not new_password:
                new_password = original_password

            if update_user(original_username, new_username, new_password, new_phone, new_role):
                messagebox.showinfo("成功", "用户修改成功！")
                dialog.destroy()
                self.reload()
            else:
                messagebox.showerror("错误", "修改失败！")

        btn_frame = tk.Frame(dialog)
        btn_frame.pack(pady=20)
        ttk.Button(btn_frame, text="确定", command=save).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="取消", command=dialog.destroy).pack(side=tk.LEFT, padx=10)

    def delete_user(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要删除的用户")
            return

        username = self.tree.item(selected[0], "values")[0]
        if not messagebox.askyesno("确认删除", f"确定要删除用户 '{username}' 吗？"):
            return

        if delete_user(username):
            messagebox.showinfo("成功", "用户删除成功！")
            self.reload()
        else:
            messagebox.showerror("错误", "删除失败！")


class BorrowManageFrame(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        tk.Label(self, text="借还书管理（BorrowManageFrame）", font=(None, 16, "bold")).pack(pady=10)

        # 使用 PanedWindow 来分割左右两部分
        paned = tk.PanedWindow(self, orient=tk.HORIZONTAL, sashwidth=8, sashrelief=tk.RAISED)
        paned.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        # === 左半部分：图书列表和借书功能 ===
        left_frame = tk.Frame(paned)
        paned.add(left_frame, stretch="always")

        tk.Label(left_frame, text="📚 图书列表（借书区）", font=(None, 12, "bold")).pack(pady=8, padx=8)

        left_btn_frame = tk.Frame(left_frame)
        left_btn_frame.pack(fill=tk.X, pady=5, padx=8)

        self.borrow_btn = ttk.Button(left_btn_frame, text="借给读者", command=self.borrow_book_dialog)
        self.borrow_btn.pack(side=tk.LEFT, padx=4)

        self.refresh_books_btn = ttk.Button(left_btn_frame, text="刷新图书", command=self.reload_books)
        self.refresh_books_btn.pack(side=tk.LEFT, padx=4)

        # 图书表格（带滚动条）
        left_scroll_frame = tk.Frame(left_frame)
        left_scroll_frame.pack(fill=tk.BOTH, expand=True, padx=8, pady=5)
        
        book_columns = ("bookname", "author", "price", "pubcom", "stock", "status")
        self.book_tree = ttk.Treeview(left_scroll_frame, columns=book_columns, show="headings")
        self.book_tree.heading("bookname", text="书名")
        self.book_tree.heading("author", text="作者")
        self.book_tree.heading("price", text="价格")
        self.book_tree.heading("pubcom", text="出版社")
        self.book_tree.heading("stock", text="库存")
        self.book_tree.heading("status", text="状态")
        self.book_tree.column("bookname", width=140, anchor=tk.W)
        self.book_tree.column("author", width=100, anchor=tk.W)
        self.book_tree.column("price", width=70, anchor=tk.CENTER)
        self.book_tree.column("pubcom", width=120, anchor=tk.W)
        self.book_tree.column("stock", width=50, anchor=tk.CENTER)
        self.book_tree.column("status", width=60, anchor=tk.CENTER)
        
        book_scroll = ttk.Scrollbar(left_scroll_frame, orient="vertical", command=self.book_tree.yview)
        self.book_tree.configure(yscrollcommand=book_scroll.set)
        
        self.book_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        book_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # === 右半部分：借阅记录和还书功能 ===
        right_frame = tk.Frame(paned)
        paned.add(right_frame, stretch="always")

        tk.Label(right_frame, text="📋 借阅记录（还书区）", font=(None, 12, "bold")).pack(pady=8, padx=8)

        right_btn_frame = tk.Frame(right_frame)
        right_btn_frame.pack(fill=tk.X, pady=5, padx=8)

        self.return_btn = ttk.Button(right_btn_frame, text="归还图书", command=self.return_book_action)
        self.return_btn.pack(side=tk.LEFT, padx=4)

        self.refresh_records_btn = ttk.Button(right_btn_frame, text="刷新记录", command=self.reload_records)
        self.refresh_records_btn.pack(side=tk.LEFT, padx=4)

        # 借阅记录表格（带滚动条）
        right_scroll_frame = tk.Frame(right_frame)
        right_scroll_frame.pack(fill=tk.BOTH, expand=True, padx=8, pady=5)
        
        record_columns = ("record_id", "username", "bookname", "borrow_date", "status")
        self.record_tree = ttk.Treeview(right_scroll_frame, columns=record_columns, show="headings")
        self.record_tree.heading("record_id", text="记录ID")
        self.record_tree.heading("username", text="读者用户名")
        self.record_tree.heading("bookname", text="图书名")
        self.record_tree.heading("borrow_date", text="借阅时间")
        self.record_tree.heading("status", text="状态")
        self.record_tree.column("record_id", width=70, anchor=tk.CENTER)
        self.record_tree.column("username", width=120, anchor=tk.W)
        self.record_tree.column("bookname", width=150, anchor=tk.W)
        self.record_tree.column("borrow_date", width=160, anchor=tk.CENTER)
        self.record_tree.column("status", width=80, anchor=tk.CENTER)
        
        record_scroll = ttk.Scrollbar(right_scroll_frame, orient="vertical", command=self.record_tree.yview)
        self.record_tree.configure(yscrollcommand=record_scroll.set)
        
        self.record_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        record_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # 初始加载数据
        self.reload_books()
        self.reload_records()

    def reload_books(self):
        """刷新图书列表。"""
        for item in self.book_tree.get_children():
            self.book_tree.delete(item)

        conn = sqlite3.connect(_book_db_path())
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT bookname, author, price, pubcom, stock, status FROM book")
            for row in cursor.fetchall():
                self.book_tree.insert("", tk.END, values=row)
        finally:
            conn.close()

    def reload_records(self):
        """刷新借阅记录列表。"""
        for item in self.record_tree.get_children():
            self.record_tree.delete(item)

        records = get_borrowing_records()
        for record in records:
            self.record_tree.insert("", tk.END, values=record)

    def borrow_book_dialog(self):
        """借书对话框。"""
        selected = self.book_tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要借出的图书")
            return

        bookname = self.book_tree.item(selected[0], "values")[0]
        
        # 创建对话框
        dialog = tk.Toplevel(self)
        dialog.title("借书")
        dialog.geometry("400x250")
        dialog.resizable(False, False)
        dialog.transient(self)
        dialog.grab_set()

        tk.Label(dialog, text=f"图书：{bookname}", font=(None, 11)).pack(pady=(20, 10))
        
        tk.Label(dialog, text="读者用户名", font=(None, 10)).pack(pady=5)
        username_entry = ttk.Entry(dialog, width=30)
        username_entry.pack(pady=5)

        # 提供现有用户作为提示
        users = load_all_users()
        user_list = [user[0] for user in users if user[3] == "student"]
        if user_list:
            tk.Label(dialog, text=f"（现有读者：{', '.join(user_list[:5])}{'...' if len(user_list) > 5 else ''}）", 
                     font=(None, 8), fg="gray").pack(pady=5)

        # 确保对话框显示后再设置焦点
        dialog.update()
        username_entry.focus_force()

        def confirm():
            username = username_entry.get().strip()
            if not username:
                messagebox.showwarning("提示", "请输入读者用户名")
                username_entry.focus_force()
                return
            
            success, msg = borrow_book(username, bookname)
            if success:
                messagebox.showinfo("成功", msg)
                dialog.destroy()
                self.reload_books()
                self.reload_records()
            else:
                messagebox.showerror("失败", msg)

        btn_frame = tk.Frame(dialog)
        btn_frame.pack(pady=20)
        ttk.Button(btn_frame, text="确认借书", command=confirm).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="取消", command=dialog.destroy).pack(side=tk.LEFT, padx=10)

    def return_book_action(self):
        """还书操作。"""
        selected = self.record_tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要归还的借阅记录")
            return

        record_id = self.record_tree.item(selected[0], "values")[0]
        bookname = self.record_tree.item(selected[0], "values")[2]
        username = self.record_tree.item(selected[0], "values")[1]

        if not messagebox.askyesno("确认还书", f"确定要让 '{username}' 归还 '{bookname}' 吗？"):
            return

        success, msg = return_book(record_id)
        if success:
            messagebox.showinfo("成功", msg)
            self.reload_books()
            self.reload_records()
        else:
            messagebox.showerror("失败", msg)


class PersonalCenterFrame(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        tk.Label(self, text="个人中心（PersonalCenterFrame）", font=(None, 16, "bold")).pack(pady=10)

        # 用户信息展示区域
        info_frame = tk.Frame(self)
        info_frame.pack(fill=tk.X, padx=20, pady=10)

        self.username_label = tk.Label(info_frame, text="用户名：", font=(None, 11))
        self.username_label.pack(anchor=tk.W, pady=5)

        self.phone_label = tk.Label(info_frame, text="手机号：", font=(None, 11))
        self.phone_label.pack(anchor=tk.W, pady=5)

        # 修改个人信息按钮
        btn_frame = tk.Frame(self)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="修改个人信息", command=self.change_info_dialog).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frame, text="修改密码", command=self.change_password_dialog).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frame, text="注销账号", command=self.delete_account).pack(side=tk.LEFT, padx=4)

        tk.Label(self, text="借阅历史", font=(None, 12, "bold")).pack(pady=5)

        # 借阅历史表格
        columns = ("bookname", "borrow_date", "return_date", "status")
        self.tree = ttk.Treeview(self, columns=columns, show="headings", height=15)
        self.tree.heading("bookname", text="书名")
        self.tree.heading("borrow_date", text="借阅日期")
        self.tree.heading("return_date", text="归还日期")
        self.tree.heading("status", text="状态")
        self.tree.column("bookname", width=200, anchor=tk.CENTER)
        self.tree.column("borrow_date", width=180, anchor=tk.CENTER)
        self.tree.column("return_date", width=180, anchor=tk.CENTER)
        self.tree.column("status", width=100, anchor=tk.CENTER)
        self.tree.pack(fill=tk.BOTH, expand=True, padx=12, pady=8)

        # 刷新按钮
        refresh_frame = tk.Frame(self)
        refresh_frame.pack(pady=5)
        ttk.Button(refresh_frame, text="刷新借阅历史", command=self.reload_records).pack()

    def load_user_info(self):
        """加载用户信息。"""
        username = self.master.current_user
        users = load_all_users()
        for user in users:
            if user[0] == username:
                self.username_label.config(text=f"用户名：{user[0]}")
                self.phone_label.config(text=f"手机号：{user[2]}")
                break

    def reload_records(self):
        """刷新借阅历史。"""
        for item in self.tree.get_children():
            self.tree.delete(item)

        records = get_user_borrow_records(self.master.current_user)
        for record in records:
            bookname, borrow_date, return_date, status = record
            return_date_display = return_date if return_date else "未归还"
            self.tree.insert("", tk.END, values=(bookname, borrow_date, return_date_display, status))

    def change_info_dialog(self):
        """修改个人信息对话框（仅修改手机号）。"""
        # 获取当前用户信息
        username = self.master.current_user
        users = load_all_users()
        current_phone = ""
        for user in users:
            if user[0] == username:
                current_phone = user[2]
                break

        # 创建对话框
        dialog = tk.Toplevel(self)
        dialog.title("修改个人信息")
        dialog.geometry("400x250")
        dialog.resizable(False, False)
        dialog.transient(self)
        dialog.grab_set()

        tk.Label(dialog, text="用户名", font=(None, 10)).pack(pady=(20, 5))
        username_entry = ttk.Entry(dialog, width=30)
        username_entry.pack(pady=5)
        username_entry.insert(0, username)
        username_entry.config(state="readonly")

        tk.Label(dialog, text="手机号", font=(None, 10)).pack(pady=5)
        phone_entry = ttk.Entry(dialog, width=30)
        phone_entry.pack(pady=5)
        phone_entry.insert(0, current_phone)

        # 确保对话框显示后再设置焦点
        dialog.update()
        phone_entry.focus_force()

        def confirm():
            new_phone = phone_entry.get().strip()

            if not new_phone:
                messagebox.showwarning("提示", "请输入手机号")
                phone_entry.focus_force()
                return

            if update_user_phone(username, new_phone):
                messagebox.showinfo("成功", "个人信息修改成功！")
                dialog.destroy()
                self.load_user_info()
            else:
                messagebox.showerror("错误", "个人信息修改失败！")

        btn_frame = tk.Frame(dialog)
        btn_frame.pack(pady=20)
        ttk.Button(btn_frame, text="确认", command=confirm).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="取消", command=dialog.destroy).pack(side=tk.LEFT, padx=10)

    def change_password_dialog(self):
        """修改密码对话框。"""
        # 获取当前用户信息
        username = self.master.current_user
        users = load_all_users()
        current_password = ""
        for user in users:
            if user[0] == username:
                current_password = user[1]
                break

        # 创建对话框
        dialog = tk.Toplevel(self)
        dialog.title("修改密码")
        dialog.geometry("400x300")
        dialog.resizable(False, False)
        dialog.transient(self)
        dialog.grab_set()

        tk.Label(dialog, text="原密码", font=(None, 10)).pack(pady=(20, 5))
        old_password_entry = ttk.Entry(dialog, width=30, show="*")
        old_password_entry.pack(pady=5)

        tk.Label(dialog, text="新密码", font=(None, 10)).pack(pady=5)
        new_password_entry = ttk.Entry(dialog, width=30, show="*")
        new_password_entry.pack(pady=5)

        tk.Label(dialog, text="确认新密码", font=(None, 10)).pack(pady=5)
        confirm_password_entry = ttk.Entry(dialog, width=30, show="*")
        confirm_password_entry.pack(pady=5)

        # 确保对话框显示后再设置焦点
        dialog.update()
        old_password_entry.focus_force()

        def confirm():
            old_pwd = old_password_entry.get().strip()
            new_pwd = new_password_entry.get().strip()
            confirm_pwd = confirm_password_entry.get().strip()

            if not old_pwd:
                messagebox.showwarning("提示", "请输入原密码")
                old_password_entry.focus_force()
                return
            if not new_pwd:
                messagebox.showwarning("提示", "请输入新密码")
                new_password_entry.focus_force()
                return
            if not confirm_pwd:
                messagebox.showwarning("提示", "请确认新密码")
                confirm_password_entry.focus_force()
                return
            if old_pwd != current_password:
                messagebox.showerror("错误", "原密码不正确")
                old_password_entry.focus_force()
                return
            if new_pwd != confirm_pwd:
                messagebox.showerror("错误", "两次输入的新密码不一致")
                new_password_entry.focus_force()
                return

            if update_user_password(username, new_pwd):
                messagebox.showinfo("成功", "密码修改成功！")
                dialog.destroy()
            else:
                messagebox.showerror("错误", "密码修改失败！")

        btn_frame = tk.Frame(dialog)
        btn_frame.pack(pady=20)
        ttk.Button(btn_frame, text="确认", command=confirm).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="取消", command=dialog.destroy).pack(side=tk.LEFT, padx=10)

    def delete_account(self):
        """注销账号。"""
        username = self.master.current_user
        if not messagebox.askyesno("警告", f"确定要注销账号 '{username}' 吗？\n此操作不可恢复！"):
            return

        if delete_user(username):
            messagebox.showinfo("成功", "账号已注销！")
            # 销毁当前管理窗口
            self.master.destroy()
        else:
            messagebox.showerror("错误", "账号注销失败！")


class ImportExportRequestFrame(tk.Frame):
    """导入导出申请处理界面（仅管理员）。"""
    def __init__(self, master=None):
        super().__init__(master)
        tk.Label(self, text="导入导出申请处理", font=(None, 16, "bold")).pack(pady=10)

        # 操作区
        btn_frame = tk.Frame(self)
        btn_frame.pack(fill=tk.X, padx=12, pady=10)

        self.approve_btn = ttk.Button(btn_frame, text="批准选中", command=self.approve_request)
        self.approve_btn.pack(side=tk.LEFT, padx=4)

        self.reject_btn = ttk.Button(btn_frame, text="拒绝选中", command=self.reject_request)
        self.reject_btn.pack(side=tk.LEFT, padx=4)

        self.refresh_btn = ttk.Button(btn_frame, text="刷新列表", command=self.reload)
        self.refresh_btn.pack(side=tk.LEFT, padx=4)

        self.select_all_btn = ttk.Button(btn_frame, text="全选", command=self.select_all)
        self.select_all_btn.pack(side=tk.LEFT, padx=4)

        self.deselect_all_btn = ttk.Button(btn_frame, text="取消全选", command=self.deselect_all)
        self.deselect_all_btn.pack(side=tk.LEFT, padx=4)

        # Treeview展示区
        columns = ("select", "request_id", "username", "request_type", "request_date", "status", "process_date", "remark")
        self.tree = ttk.Treeview(self, columns=columns, show="headings", height=18)
        self.tree.heading("select", text="选择")
        self.tree.heading("request_id", text="申请ID")
        self.tree.heading("username", text="申请人")
        self.tree.heading("request_type", text="申请类型")
        self.tree.heading("request_date", text="申请时间")
        self.tree.heading("status", text="状态")
        self.tree.heading("process_date", text="处理时间")
        self.tree.heading("remark", text="备注")
        self.tree.column("select", width=50, anchor=tk.CENTER)
        self.tree.column("request_id", width=60, anchor=tk.CENTER)
        self.tree.column("username", width=100, anchor=tk.CENTER)
        self.tree.column("request_type", width=80, anchor=tk.CENTER)
        self.tree.column("request_date", width=140, anchor=tk.CENTER)
        self.tree.column("status", width=80, anchor=tk.CENTER)
        self.tree.column("process_date", width=140, anchor=tk.CENTER)
        self.tree.column("remark", width=150, anchor=tk.W)
        
        # 滚动条
        scroll = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scroll.set)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=12, pady=8)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 存储选中状态
        self.selected_items = set()
        
        # 绑定点击事件切换选中
        self.tree.bind("<Button-1>", self.on_click)

        # 初始加载数据
        self.reload()

    def on_click(self, event):
        """点击事件 - 切换选中状态"""
        item = self.tree.identify_row(event.y)
        if not item:
            return

        # 切换选中状态
        if item in self.selected_items:
            self.selected_items.remove(item)
            self.tree.set(item, "select", "")
        else:
            self.selected_items.add(item)
            self.tree.set(item, "select", "☑")

    def select_all(self):
        """全选"""
        self.selected_items.clear()
        for item in self.tree.get_children():
            self.selected_items.add(item)
            self.tree.set(item, "select", "☑")

    def deselect_all(self):
        """取消全选"""
        for item in self.tree.get_children():
            self.tree.set(item, "select", "")
        self.selected_items.clear()

    def reload(self):
        """刷新申请列表。"""
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.selected_items.clear()

        requests = get_all_import_export_requests()
        for req in requests:
            request_id, username, request_type, request_date, status, process_date, remark = req
            process_date_display = process_date if process_date else "-"
            remark_display = remark if remark else "-"
            self.tree.insert("", tk.END, values=("", request_id, username, request_type, request_date, status, process_date_display, remark_display))

    def approve_request(self):
        """批量批准申请。"""
        if not self.selected_items:
            messagebox.showwarning("提示", "请先选择要批准的申请")
            return

        # 获取选中的待处理申请
        selected_requests = []
        for item in self.selected_items:
            values = self.tree.item(item, "values")
            request_id = values[1]
            status = values[5]
            if status == "待处理":
                selected_requests.append(request_id)

        if not selected_requests:
            messagebox.showwarning("提示", "选中的申请都已被处理")
            return

        # 创建对话框询问备注
        dialog = tk.Toplevel(self)
        dialog.title("批量批准申请")
        dialog.geometry("400x250")
        dialog.resizable(False, False)
        dialog.transient(self)
        dialog.grab_set()

        tk.Label(dialog, text="备注（可选）", font=(None, 10)).pack(pady=(30, 5))
        remark_entry = ttk.Entry(dialog, width=35)
        remark_entry.pack(pady=5)

        def confirm():
            remark = remark_entry.get().strip()
            success_count = 0
            fail_count = 0
            
            for request_id in selected_requests:
                if update_import_export_request_status(request_id, "已批准", remark):
                    success_count += 1
                else:
                    fail_count += 1
            
            messagebox.showinfo("结果", f"处理完成！成功 {success_count} 个，失败 {fail_count} 个")
            dialog.destroy()
            self.reload()

        btn_frame = tk.Frame(dialog)
        btn_frame.pack(pady=30)
        ttk.Button(btn_frame, text="确认", command=confirm).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="取消", command=dialog.destroy).pack(side=tk.LEFT, padx=10)

    def reject_request(self):
        """批量拒绝申请。"""
        if not self.selected_items:
            messagebox.showwarning("提示", "请先选择要拒绝的申请")
            return

        # 获取选中的待处理申请
        selected_requests = []
        for item in self.selected_items:
            values = self.tree.item(item, "values")
            request_id = values[1]
            status = values[5]
            if status == "待处理":
                selected_requests.append(request_id)

        if not selected_requests:
            messagebox.showwarning("提示", "选中的申请都已被处理")
            return

        # 创建对话框询问备注
        dialog = tk.Toplevel(self)
        dialog.title("批量拒绝申请")
        dialog.geometry("400x250")
        dialog.resizable(False, False)
        dialog.transient(self)
        dialog.grab_set()

        tk.Label(dialog, text="备注（可选）", font=(None, 10)).pack(pady=(30, 5))
        remark_entry = ttk.Entry(dialog, width=35)
        remark_entry.pack(pady=5)

        def confirm():
            remark = remark_entry.get().strip()
            success_count = 0
            fail_count = 0
            
            for request_id in selected_requests:
                if update_import_export_request_status(request_id, "已拒绝", remark):
                    success_count += 1
                else:
                    fail_count += 1
            
            messagebox.showinfo("结果", f"处理完成！成功 {success_count} 个，失败 {fail_count} 个")
            dialog.destroy()
            self.reload()

        btn_frame = tk.Frame(dialog)
        btn_frame.pack(pady=30)
        ttk.Button(btn_frame, text="确认", command=confirm).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="取消", command=dialog.destroy).pack(side=tk.LEFT, padx=10)


class StudentImportExportFrame(tk.Frame):
    """学生导入导出申请页面。"""
    def __init__(self, master=None):
        super().__init__(master)
        tk.Label(self, text="导入导出申请", font=(None, 16, "bold")).pack(pady=10)

        # 操作区
        btn_frame = tk.Frame(self)
        btn_frame.pack(fill=tk.X, padx=12, pady=10)

        self.request_import_btn = ttk.Button(btn_frame, text="申请导入", command=self.request_import)
        self.request_import_btn.pack(side=tk.LEFT, padx=4)

        self.request_export_btn = ttk.Button(btn_frame, text="申请导出", command=self.request_export)
        self.request_export_btn.pack(side=tk.LEFT, padx=4)

        self.refresh_btn = ttk.Button(btn_frame, text="刷新状态", command=self.reload)
        self.refresh_btn.pack(side=tk.LEFT, padx=4)

        self.select_all_btn = ttk.Button(btn_frame, text="全选", command=self.select_all)
        self.select_all_btn.pack(side=tk.LEFT, padx=4)

        self.deselect_all_btn = ttk.Button(btn_frame, text="取消全选", command=self.deselect_all)
        self.deselect_all_btn.pack(side=tk.LEFT, padx=4)

        # 导入导出按钮（只在有已批准申请时显示）
        self.action_btn_frame = tk.Frame(self)
        self.action_btn_frame.pack(fill=tk.X, padx=12, pady=5)

        self.import_btn = ttk.Button(self.action_btn_frame, text="执行导入", command=self.do_import)
        self.import_btn.pack(side=tk.LEFT, padx=4)
        self.import_btn.pack_forget()  # 初始隐藏

        self.export_btn = ttk.Button(self.action_btn_frame, text="执行导出", command=self.do_export)
        self.export_btn.pack(side=tk.LEFT, padx=4)
        self.export_btn.pack_forget()  # 初始隐藏

        # 提示标签
        self.hint_label = tk.Label(self, text="", font=(None, 10), fg="blue")
        self.hint_label.pack(pady=5)

        # Treeview展示区
        columns = ("select", "request_id", "request_type", "request_date", "status", "process_date", "remark")
        self.tree = ttk.Treeview(self, columns=columns, show="headings", height=15)
        self.tree.heading("select", text="选择")
        self.tree.heading("request_id", text="申请ID")
        self.tree.heading("request_type", text="申请类型")
        self.tree.heading("request_date", text="申请时间")
        self.tree.heading("status", text="状态")
        self.tree.heading("process_date", text="处理时间")
        self.tree.heading("remark", text="备注")
        self.tree.column("select", width=50, anchor=tk.CENTER)
        self.tree.column("request_id", width=60, anchor=tk.CENTER)
        self.tree.column("request_type", width=80, anchor=tk.CENTER)
        self.tree.column("request_date", width=140, anchor=tk.CENTER)
        self.tree.column("status", width=80, anchor=tk.CENTER)
        self.tree.column("process_date", width=140, anchor=tk.CENTER)
        self.tree.column("remark", width=200, anchor=tk.W)
        
        # 滚动条
        scroll = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scroll.set)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=12, pady=8)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 存储选中状态
        self.selected_items = set()
        
        # 绑定点击事件切换选中
        self.tree.bind("<Button-1>", self.on_click)

        # 初始加载数据
        self.reload()

    def on_click(self, event):
        """点击事件 - 切换选中状态"""
        item = self.tree.identify_row(event.y)
        if not item:
            return

        # 切换选中状态
        if item in self.selected_items:
            self.selected_items.remove(item)
            self.tree.set(item, "select", "")
        else:
            self.selected_items.add(item)
            self.tree.set(item, "select", "☑")

    def select_all(self):
        """全选"""
        self.selected_items.clear()
        for item in self.tree.get_children():
            self.selected_items.add(item)
            self.tree.set(item, "select", "☑")

    def deselect_all(self):
        """取消全选"""
        for item in self.tree.get_children():
            self.tree.set(item, "select", "")
        self.selected_items.clear()

    def reload(self):
        """刷新申请列表和按钮状态。"""
        # 清空Treeview
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.selected_items.clear()

        # 获取当前用户的申请
        requests = get_user_import_export_requests(self.master.current_user)
        has_approved_import = False
        has_approved_export = False

        for req in requests:
            request_id, request_type, request_date, status, process_date, remark = req
            process_date_display = process_date if process_date else "-"
            remark_display = remark if remark else "-"
            self.tree.insert("", tk.END, values=("", request_id, request_type, request_date, status, process_date_display, remark_display))

            # 检查是否有已批准的申请
            if status == "已批准":
                if request_type == "导入":
                    has_approved_import = True
                elif request_type == "导出":
                    has_approved_export = True

        # 更新按钮显示
        if has_approved_import:
            self.import_btn.pack(side=tk.LEFT, padx=4)
        else:
            self.import_btn.pack_forget()

        if has_approved_export:
            self.export_btn.pack(side=tk.LEFT, padx=4)
        else:
            self.export_btn.pack_forget()

        # 更新提示信息
        if has_approved_import or has_approved_export:
            hints = []
            if has_approved_import:
                hints.append("导入")
            if has_approved_export:
                hints.append("导出")
            self.hint_label.config(text=f"✓ 您有已批准的{ '、'.join(hints) }申请，可以执行相应操作！")
        else:
            self.hint_label.config(text="")

    def request_import(self):
        """申请导入。"""
        if add_import_export_request(self.master.current_user, "导入"):
            messagebox.showinfo("成功", "导入申请已提交，请等待管理员批准！")
            self.reload()
        else:
            messagebox.showerror("错误", "申请提交失败！")

    def request_export(self):
        """申请导出。"""
        if add_import_export_request(self.master.current_user, "导出"):
            messagebox.showinfo("成功", "导出申请已提交，请等待管理员批准！")
            self.reload()
        else:
            messagebox.showerror("错误", "申请提交失败！")

    def do_import(self):
        """执行导入。"""
        # 直接调用ManageWin的import_data方法
        self.master.import_data()
        self.reload()

    def do_export(self):
        """执行导出。"""
        # 直接调用ManageWin的export_data方法
        self.master.export_data()
        self.reload()


class DoubanSpiderFrame(tk.Frame):
    """豆瓣图书爬虫集成界面（仅管理员可见）。"""
    
    def __init__(self, master=None):
        super().__init__(master)
        tk.Label(self, text="豆瓣图书爬虫", font=(None, 16, "bold")).pack(pady=10)
        
        try:
            # 导入爬虫模块
            from douban_spider import DoubanSpider
            self.spider = DoubanSpider()
            self.spider_initialized = True
        except Exception as e:
            self.spider_initialized = False
            tk.Label(self, text=f"⚠️  爬虫模块初始化失败: {e}", fg="red").pack(pady=10)
        
        # 标签区域
        tag_frame = tk.LabelFrame(self, text="第一步：选择图书分类")
        tag_frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=8)
        
        # 标签选择Treeview（带滚动条和复选框）
        tree_container = tk.Frame(tag_frame)
        tree_container.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        
        tag_columns = ("name", "url")
        self.tag_tree = ttk.Treeview(tree_container, columns=tag_columns, show="tree headings", height=8)
        self.tag_tree.heading("#0", text="选择")
        self.tag_tree.heading("name", text="标签名称")
        self.tag_tree.heading("url", text="URL")
        self.tag_tree.column("#0", width=40, anchor=tk.CENTER)
        self.tag_tree.column("name", width=150, anchor=tk.W)
        self.tag_tree.column("url", width=400, anchor=tk.W)
        
        # 滚动条
        tree_scrollbar = ttk.Scrollbar(tree_container, orient="vertical", command=self.tag_tree.yview)
        self.tag_tree.configure(yscrollcommand=tree_scrollbar.set)
        
        self.tag_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tree_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 存储选中状态
        self.tag_selected = set()
        
        # 绑定点击事件切换选中
        self.tag_tree.bind("<Button-1>", self.on_tag_click)
        
        # 选择提示标签
        self.selection_hint = tk.Label(tag_frame, text="💡 提示：点击左侧方框选择图书分类，可多选", fg="#666", font=(None, 9))
        self.selection_hint.pack(pady=(0, 5))
        
        # 标签操作按钮
        tag_btn_frame = tk.Frame(tag_frame)
        tag_btn_frame.pack(fill=tk.X, padx=8, pady=5)
        ttk.Button(tag_btn_frame, text="刷新标签", command=self.refresh_tags).pack(side=tk.LEFT, padx=4)
        ttk.Button(tag_btn_frame, text="全选", command=self.select_all_tags).pack(side=tk.LEFT, padx=4)
        ttk.Button(tag_btn_frame, text="取消全选", command=self.deselect_all_tags).pack(side=tk.LEFT, padx=4)
        ttk.Button(tag_btn_frame, text="确认选择", command=self.confirm_selection).pack(side=tk.LEFT, padx=4)
        
        # 爬取设置
        setting_frame = tk.LabelFrame(self, text="第二步：设置爬取参数")
        setting_frame.pack(fill=tk.X, padx=12, pady=8)
        
        tk.Label(setting_frame, text="每类图书爬取页数：").pack(side=tk.LEFT, padx=12, pady=10)
        self.page_count_var = tk.IntVar(value=1)
        ttk.Spinbox(setting_frame, from_=1, to=100, textvariable=self.page_count_var, width=8).pack(side=tk.LEFT, padx=4, pady=10)
        tk.Label(setting_frame, text="（建议1-5页，每页约20本）", fg="#666", font=(None, 9)).pack(side=tk.LEFT, padx=4, pady=10)
        
        self.skip_duplicate_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(setting_frame, text="跳过已存在的图书", variable=self.skip_duplicate_var).pack(side=tk.LEFT, padx=12, pady=10)
        
        # 爬取操作
        crawl_frame = tk.Frame(self)
        crawl_frame.pack(fill=tk.X, padx=12, pady=8)
        
        self.start_btn = ttk.Button(crawl_frame, text="开始爬取", command=self.start_crawl, width=15)
        self.start_btn.pack(side=tk.LEFT, padx=4)
        
        self.stop_btn = ttk.Button(crawl_frame, text="停止爬取", command=self.stop_crawl, width=15, state="disabled")
        self.stop_btn.pack(side=tk.LEFT, padx=4)
        
        ttk.Button(crawl_frame, text="清理缓存", command=self.clear_cache).pack(side=tk.LEFT, padx=4)
        
        # 进度和日志
        log_frame = tk.LabelFrame(self, text="第三步：查看进度和结果")
        log_frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=8)
        
        self.log_text = tk.Text(log_frame, height=10, state="disabled")
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        
        # 数据预览和导入
        result_frame = tk.LabelFrame(self, text="第四步：预览和导入数据")
        result_frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=8)
        
        preview_columns = ("bookname", "author", "price", "pubcom", "stock")
        self.preview_tree = ttk.Treeview(result_frame, columns=preview_columns, show="headings", height=5)
        self.preview_tree.heading("bookname", text="书名")
        self.preview_tree.heading("author", text="作者")
        self.preview_tree.heading("price", text="价格")
        self.preview_tree.heading("pubcom", text="出版社")
        self.preview_tree.heading("stock", text="库存")
        self.preview_tree.column("bookname", width=200, anchor=tk.W)
        self.preview_tree.column("author", width=120, anchor=tk.W)
        self.preview_tree.column("price", width=80, anchor=tk.CENTER)
        self.preview_tree.column("pubcom", width=150, anchor=tk.W)
        self.preview_tree.column("stock", width=60, anchor=tk.CENTER)
        self.preview_tree.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        
        result_btn_frame = tk.Frame(result_frame)
        result_btn_frame.pack(fill=tk.X, padx=8, pady=5)
        ttk.Button(result_btn_frame, text="导入到系统", command=self.import_to_system).pack(side=tk.LEFT, padx=4)
        ttk.Button(result_btn_frame, text="导出为CSV", command=self.export_csv).pack(side=tk.LEFT, padx=4)
        
        self.crawling = False
        self.crawled_books = []
    
    def log(self, message, level="info"):
        """添加日志消息"""
        from datetime import datetime
        timestamp = datetime.now().strftime("%H:%M:%S")
        
        self.log_text.config(state="normal")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)
        self.log_text.config(state="disabled")
    
    def on_tag_click(self, event):
        """标签点击事件 - 切换选中状态"""
        item = self.tag_tree.identify_row(event.y)
        if not item:
            return
        
        # 切换选中状态
        if item in self.tag_selected:
            self.tag_selected.remove(item)
            self.tag_tree.item(item, text="")
        else:
            self.tag_selected.add(item)
            self.tag_tree.item(item, text="☑")
        
        # 更新提示
        count = len(self.tag_selected)
        if count > 0:
            self.selection_hint.config(text=f"已选择 {count} 个分类", fg="#006600")
        else:
            self.selection_hint.config(text="💡 提示：点击左侧方框选择图书分类，可多选", fg="#666")
    
    def refresh_tags(self):
        """刷新标签列表"""
        if not self.spider_initialized:
            messagebox.showerror("错误", "爬虫模块未初始化")
            return
        
        self.log("正在加载标签列表...")
        try:
            # 清空旧数据
            for item in self.tag_tree.get_children():
                self.tag_tree.delete(item)
            self.tag_selected.clear()
            
            # 加载标签
            tags = self.spider.load_tags()
            
            for tag in tags:
                self.tag_tree.insert("", tk.END, values=(tag.name, tag.url), text="")
            
            self.log(f"✅ 已加载 {len(tags)} 个图书分类")
            self.selection_hint.config(text=f"💡 提示：共 {len(tags)} 个分类，请选择要爬取的图书分类", fg="#666")
        except Exception as e:
            self.log(f"加载标签失败: {e}", level="error")
            messagebox.showerror("错误", f"加载标签失败: {e}")
    
    def select_all_tags(self):
        """全选标签"""
        self.tag_selected.clear()
        for item in self.tag_tree.get_children():
            self.tag_selected.add(item)
            self.tag_tree.item(item, text="☑")
        self.selection_hint.config(text=f"已选择 {len(self.tag_selected)} 个分类", fg="#006600")
    
    def deselect_all_tags(self):
        """取消全选标签"""
        for item in self.tag_tree.get_children():
            self.tag_tree.item(item, text="")
        self.tag_selected.clear()
        self.selection_hint.config(text="💡 提示：点击左侧方框选择图书分类，可多选", fg="#666")
    
    def confirm_selection(self):
        """确认标签选择"""
        if not self.tag_selected:
            messagebox.showwarning("提示", "请先选择要爬取的图书分类")
            return
        
        tag_names = [self.tag_tree.item(item, "values")[0] for item in self.tag_selected]
        
        # 显示确认对话框
        tag_list_text = "\n".join([f"  • {name}" for name in tag_names[:10]])
        if len(tag_names) > 10:
            tag_list_text += f"\n  ... 还有 {len(tag_names) - 10} 个分类"
        
        confirm_msg = (
            f"已选择 {len(tag_names)} 个图书分类：\n\n"
            f"{tag_list_text}\n\n"
            f"确定要爬取这些分类吗？"
        )
        
        if messagebox.askyesno("确认选择", confirm_msg):
            self.selection_hint.config(
                text=f"✅ 已选择 {len(tag_names)} 个分类，准备就绪！",
                fg="#006600"
            )
            self.log(f"✅ 已选择 {len(tag_names)} 个图书分类")
    
    def start_crawl(self):
        """开始爬取"""
        if not self.spider_initialized:
            messagebox.showerror("错误", "爬虫模块未初始化")
            return
        
        # 获取选中的标签
        if not self.tag_selected:
            messagebox.showwarning("提示", "请先选择要爬取的图书分类")
            return
        
        tag_names = [self.tag_tree.item(item, "values")[0] for item in self.tag_selected]
        self.spider.set_selected_tags(tag_names)
        
        # 获取页数（确保这里正确获取）
        page_count = int(self.page_count_var.get())
        self.log(f"📋 准备爬取 {len(tag_names)} 个分类，每类 {page_count} 页")
        
        # 清空预览
        for item in self.preview_tree.get_children():
            self.preview_tree.delete(item)
        self.crawled_books = []
        
        # 清空日志并开始爬取
        self.log_text.config(state="normal")
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state="disabled")
        
        # 设置按钮状态
        if hasattr(self, 'start_btn') and self.start_btn:
            self.start_btn.config(state="disabled")
        if hasattr(self, 'stop_btn') and self.stop_btn:
            self.stop_btn.config(state="normal")
        
        # 设置进度回调
        def progress_callback(message, level="info"):
            self.master.after(0, lambda: self.log(message, level))
        
        self.spider.set_progress_callback(progress_callback)
        
        # 开始爬取（使用线程避免界面卡顿）
        import threading
        
        self.crawling = True
        
        # 在这里保存变量值，避免闭包延迟绑定问题
        local_tag_count = len(tag_names)
        local_page_count = page_count
        
        def crawl_thread():
            try:
                self.log(f"🚀 开始爬取 {local_tag_count} 个分类，每类 {local_page_count} 页...")
                
                self.crawled_books = self.spider.crawl(page_count=local_page_count)
                
                # 更新预览
                self.master.after(0, self.update_preview)
                
                self.master.after(0, lambda: messagebox.showinfo("成功", f"爬取完成！共获取 {len(self.crawled_books)} 本图书"))
                
            except Exception as e:
                self.log(f"❌ 爬取出错: {e}", level="error")
                self.master.after(0, lambda: messagebox.showerror("错误", f"爬取出错: {e}"))
            finally:
                self.crawling = False
                # 恢复按钮状态
                self.master.after(0, self.restore_buttons)
        
        thread = threading.Thread(target=crawl_thread)
        thread.daemon = True
        thread.start()
    
    def restore_buttons(self):
        """恢复按钮状态"""
        if hasattr(self, 'start_btn') and self.start_btn:
            self.start_btn.config(state="normal")
        if hasattr(self, 'stop_btn') and self.stop_btn:
            self.stop_btn.config(state="disabled")
    
    def stop_crawl(self):
        """停止爬取"""
        if not self.crawling:
            messagebox.showinfo("提示", "当前没有正在进行的爬取任务")
            return
        
        self.crawling = False
        self.spider.stop()
        self.log("⏹ 正在停止爬取...")
    
    def update_preview(self):
        """更新预览列表"""
        for item in self.preview_tree.get_children():
            self.preview_tree.delete(item)
        
        for book in self.crawled_books[:100]:  # 只显示前100本
            self.preview_tree.insert("", tk.END, values=(
                book.bookname,
                book.author,
                book.price,
                book.pubcom,
                book.stock
            ))
    
    def import_to_system(self):
        """导入到图书管理系统"""
        if not self.crawled_books:
            messagebox.showwarning("提示", "没有可导入的数据，请先爬取图书")
            return
        
        try:
            skip_duplicate = self.skip_duplicate_var.get()
            imported_count, skipped_count = self.spider.import_to_system(skip_duplicate=skip_duplicate)
            
            self.log(f"导入完成: 成功 {imported_count} 本，跳过 {skipped_count} 本")
            messagebox.showinfo("成功", f"导入完成！\n成功导入: {imported_count} 本\n跳过重复: {skipped_count} 本")
            
            # 刷新图书列表
            if hasattr(self.master, "list_frame") and hasattr(self.master.list_frame, "reload"):
                self.master.list_frame.reload()
                
        except Exception as e:
            self.log(f"导入失败: {e}", level="error")
            messagebox.showerror("错误", f"导入失败: {e}")
    
    def export_csv(self):
        """导出为CSV"""
        if not self.crawled_books:
            messagebox.showwarning("提示", "没有可导出的数据，请先爬取图书")
            return
        
        try:
            file_path = self.spider.export_csv()
            self.log(f"已导出CSV: {file_path}")
            messagebox.showinfo("成功", f"已导出到: {file_path}")
        except Exception as e:
            self.log(f"导出失败: {e}", level="error")
            messagebox.showerror("错误", f"导出失败: {e}")
    
    def clear_cache(self):
        """清理缓存"""
        if not self.spider_initialized:
            return
        
        from tkinter.simpledialog import askstring
        result = messagebox.askyesno("确认", "确定要清理缓存文件吗？\n\n选项：\n- 是：清理列表页和详情页缓存\n- 否：仅清理详情页")
        
        if result:
            from douban_spider import clear_cache
            count = clear_cache("all")
            self.log(f"已清理 {count} 个缓存文件")
        elif messagebox.askyesno("确认", "清理详情页缓存？"):
            from douban_spider import clear_cache
            count = clear_cache("detail")
            self.log(f"已清理 {count} 个缓存文件")


class ManageWin(tk.Tk):
    def __init__(self, role='student', current_user=''):
        super().__init__()
        self.role = role
        self.current_user = current_user
        self.title('后台管理界面')
        self.geometry('800x750')
        self.resizable(True, True)

        # 初始化数据库图书表
        init_book_table()
        # 迁移 book 表字段（添加 stock 和 status）
        migrate_book_fields()
        # 初始化借阅记录表
        init_borrow_table()
        # 初始化导入导出申请表
        init_import_export_request_table()
        # 初始化借阅申请表
        init_borrow_request_table()

        # 预创建各页面
        self.welcome_frame = WelcomeFrame(self)
        self.list_frame = ListFrame(self)
        self.add_frame = AddFrame(self)
        self.data_frame = DataFrame(self)
        self.about_frame = AboutFrame(self)
        self.user_manage_frame = UserManageFrame(self)
        self.borrow_manage_frame = BorrowManageFrame(self)
        self.personal_center_frame = PersonalCenterFrame(self)
        self.import_export_request_frame = ImportExportRequestFrame(self)
        self.student_import_export_frame = StudentImportExportFrame(self)
        self.douban_spider_frame = DoubanSpiderFrame(self)
        # 借阅申请相关页面
        self.student_book_list_frame = StudentBookListFrame(self)
        self.my_borrow_requests_frame = MyBorrowRequestsFrame(self)
        self.borrow_request_approval_frame = BorrowRequestApprovalFrame(self)
        self.current_frame = None

        # 创建菜单栏
        self.create_menu()

        # 初始显示欢迎页
        self.showFrame(self.welcome_frame)

    def create_menu(self):
        menubar = tk.Menu(self)

        # 图书管理
        book_menu = tk.Menu(menubar, tearoff=0)
        book_menu.add_command(label="欢迎页", command=lambda: self.showFrame(self.welcome_frame))
        book_menu.add_command(label="图书列表", command=lambda: self.showFrame(self.list_frame))
        if self.role == "admin":
            book_menu.add_command(label="添加图书", command=lambda: self.showFrame(self.add_frame))
        menubar.add_cascade(label="图书管理", menu=book_menu)

        # 借阅申请（学生）/ 审批（管理员）
        if self.role == "student":
            borrow_menu = tk.Menu(menubar, tearoff=0)
            borrow_menu.add_command(label="图书列表", command=lambda: self.showFrame(self.student_book_list_frame))
            borrow_menu.add_command(label="我的申请", command=lambda: self.showFrame(self.my_borrow_requests_frame))
            menubar.add_cascade(label="借阅申请", menu=borrow_menu)
        else:
            menubar.add_command(label="借阅审批", command=lambda: self.showFrame(self.borrow_request_approval_frame))

        # 用户管理（仅管理员，点击直达）
        if self.role == "admin":
            menubar.add_command(label="用户管理", command=lambda: self.showFrame(self.user_manage_frame))

        # 借还书管理（仅管理员，点击直达）
        if self.role == "admin":
            menubar.add_command(label="借还书管理", command=lambda: self.showFrame(self.borrow_manage_frame))

        # 个人中心（仅学生用户，点击直达）
        if self.role == "student":
            menubar.add_command(label="个人中心", command=lambda: self.showFrame(self.personal_center_frame))

        # 数据分析
        menubar.add_command(label="数据分析", command=lambda: self.showFrame(self.data_frame))

        # 豆瓣图书爬虫（仅管理员，点击直达）
        if self.role == "admin":
            menubar.add_command(label="豆瓣爬虫", command=lambda: self.showFrame(self.douban_spider_frame))

        # 导入导出
        if self.role == "admin":
            # 管理员：二级菜单
            io_menu = tk.Menu(menubar, tearoff=0)
            io_menu.add_command(label="处理申请", command=lambda: self.showFrame(self.import_export_request_frame))
            io_menu.add_separator()
            io_menu.add_command(label="导入", command=self.import_data)
            io_menu.add_command(label="导出", command=self.export_data)
            menubar.add_cascade(label="导入导出", menu=io_menu)
        else:
            # 学生：一级菜单直达页面
            menubar.add_command(label="导入导出", command=lambda: self.showFrame(self.student_import_export_frame))

        # 帮助
        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label="关于我们", command=lambda: self.showFrame(self.about_frame))
        help_menu.add_separator()
        help_menu.add_command(label="退出系统", command=self.destroy)
        menubar.add_cascade(label="帮助", menu=help_menu)

        self.config(menu=menubar)

    def request_import_data(self):
        """学生申请导入数据。"""
        if add_import_export_request(self.current_user, "导入"):
            messagebox.showinfo("成功", "导入申请已提交，请等待管理员批准！")
        else:
            messagebox.showerror("错误", "申请提交失败！")

    def request_export_data(self):
        """学生申请导出数据。"""
        if add_import_export_request(self.current_user, "导出"):
            messagebox.showinfo("成功", "导出申请已提交，请等待管理员批准！")
        else:
            messagebox.showerror("错误", "申请提交失败！")

    def export_data(self):
        """导出数据（只有管理员或有已批准申请的学生可执行）。"""
        # 检查权限
        if self.role != "admin" and not check_import_export_approved(self.current_user, "导出"):
            messagebox.showwarning("提示", "您没有已批准的导出申请！请先申请并等待管理员批准。")
            return

        file_path = asksaveasfilename(
            title="导出图书数据",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
        )
        if not file_path:
            return

        conn = sqlite3.connect(_book_db_path())
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT bookname, author, price, pubcom, stock, status FROM book")
            rows = cursor.fetchall()
        finally:
            conn.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "图书数据"
        ws.append(["书名", "作者", "价格", "出版社", "库存", "状态"])
        for row in rows:
            ws.append(list(row))

        try:
            wb.save(file_path)
            messagebox.showinfo("提示", "导出成功")
        except PermissionError:
            messagebox.showerror("错误", "导出失败！文件正被其他程序占用，请关闭后重试。")
        except Exception as e:
            messagebox.showerror("错误", f"导出时发生未知错误：{e}")

    def import_data(self):
        """导入数据（只有管理员或有已批准申请的学生可执行）。"""
        # 检查权限
        if self.role != "admin" and not check_import_export_approved(self.current_user, "导入"):
            messagebox.showwarning("提示", "您没有已批准的导入申请！请先申请并等待管理员批准。")
            return

        file_path = askopenfilename(
            title="导入图书数据",
            filetypes=[("Excel 文件", "*.xlsx *.xlsm *.xltx *.xltm")],
        )
        if not file_path:
            return

        try:
            wb = load_workbook(file_path)
            ws = wb.active
        except Exception as e:
            messagebox.showerror("错误", f"读取 Excel 失败：{e}")
            return

        conn = sqlite3.connect(_book_db_path())
        try:
            cursor = conn.cursor()
            success_count = 0
            error_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue

                try:
                    values = list(row) + [None] * (6 - len(row))
                    bookname, author, price, pubcom, stock, status = values[0], values[1], values[2], values[3], values[4], values[5]
                    if bookname is None or str(bookname).strip() == "":
                        error_count += 1
                        continue

                    # 处理库存默认值
                    if stock is None or str(stock).strip() == "":
                        stock = 1
                    else:
                        try:
                            stock = int(stock)
                        except ValueError:
                            stock = 1

                    # 处理状态默认值
                    if status is None or str(status).strip() == "":
                        status = "在馆"
                    else:
                        status = str(status).strip()
                        if status not in ["在馆", "借出"]:
                            status = "在馆"

                    cursor.execute(
                        "INSERT OR REPLACE INTO book (bookname, price, author, pubcom, stock, status) VALUES (?, ?, ?, ?, ?, ?)",
                        (
                            str(bookname).strip(),
                            "" if price is None else str(price).strip(),
                            "" if author is None else str(author).strip(),
                            "" if pubcom is None else str(pubcom).strip(),
                            stock,
                            status,
                        ),
                    )
                    success_count += 1
                except Exception:
                    error_count += 1
            conn.commit()
        except Exception as e:
            messagebox.showerror("错误", f"导入失败：{e}")
            return
        finally:
            conn.close()

        messagebox.showinfo("提示", f"导入完成：成功 {success_count} 条，失败 {error_count} 条")
        self.list_frame.reload()

    def showFrame(self, frame):
        """
        利用 pack() 和 pack_forget() 实现页面切换
        """
        if self.current_frame is not None:
            self.current_frame.pack_forget()
        self.current_frame = frame
        self.current_frame.pack(fill=tk.BOTH, expand=True)
        
        # 如果是 AddFrame，自动聚焦到第一个输入框
        if isinstance(frame, AddFrame):
            self.after(100, lambda: frame.bookname_entry.focus_set())
        
        # 如果是 PersonalCenterFrame，加载用户信息和借阅记录
        if isinstance(frame, PersonalCenterFrame):
            frame.load_user_info()
            frame.reload_records()

        # 如果是 StudentImportExportFrame，刷新申请列表
        if isinstance(frame, StudentImportExportFrame):
            frame.reload()
            
        # 如果是 StudentBookListFrame，刷新图书列表
        if isinstance(frame, StudentBookListFrame):
            frame.reload()
            
        # 如果是 MyBorrowRequestsFrame，刷新申请列表
        if isinstance(frame, MyBorrowRequestsFrame):
            frame.reload()
            
        # 如果是 BorrowRequestApprovalFrame，刷新申请列表
        if isinstance(frame, BorrowRequestApprovalFrame):
            frame.reload()
            
        # 如果是 ImportExportRequestFrame，刷新申请列表
        if isinstance(frame, ImportExportRequestFrame):
            frame.reload()
            
        # 如果是 ListFrame，刷新图书列表
        if isinstance(frame, ListFrame):
            frame.reload()


if __name__ == "__main__":
    app = ManageWin()
    app.mainloop()
